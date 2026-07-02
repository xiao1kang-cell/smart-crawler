from __future__ import annotations

import pytest
from fastapi import HTTPException
from datetime import date, datetime, timedelta
from unittest.mock import patch
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.apikey import hash_key
from app.auth import hash_password, hash_secret, make_token
from app.db import Base
from app.models import (ApiKey, CrawlFailure, CrawlJob, CrawlUrl, InviteCode, OnDemandJob,
                        PriceHistory, Product, Promotion, Review, Site,
                        SiteMetric, SpineJob, User, Workspace, WorkspaceMember,
                        WorkspaceSite, Trend)


pytestmark = pytest.mark.unit


def _session():
    from app.api.routes import _COVERAGE_CACHE

    _COVERAGE_CACHE.clear()
    engine = create_engine("sqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    return Session()


def _threadsafe_session():
    from app.api.routes import _COVERAGE_CACHE

    _COVERAGE_CACHE.clear()
    engine = create_engine("sqlite://", future=True,
                           connect_args={"check_same_thread": False},
                           poolclass=StaticPool)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    return Session()


def _workspace(db, name: str, slug: str) -> Workspace:
    row = Workspace(name=name, slug=slug, type="customer", status="active")
    db.add(row)
    db.flush()
    return row


def _user(db, username: str, workspace: Workspace, *,
          role: str = "user", global_role: str | None = None) -> User:
    row = User(
        username=username,
        email=f"{username}@example.com",
        password_hash=hash_password("Password1"),
        role=role,
        global_role=global_role,
        status="active",
        default_workspace_id=workspace.id,
    )
    db.add(row)
    db.flush()
    db.add(WorkspaceMember(workspace_id=workspace.id, user_id=row.id,
                           role="owner" if role == "admin" else "member",
                           status="active"))
    db.flush()
    return row


def _site(db, code: str, brand: str) -> Site:
    row = Site(site=code, brand=brand, country="US",
               url=f"https://{code}.example.com", platform="generic")
    db.add(row)
    db.flush()
    return row


def _workspace_site(db, workspace: Workspace, site: str, *,
                    enabled: bool = True,
                    target_sku_count: int | None = None) -> WorkspaceSite:
    row = WorkspaceSite(workspace_id=workspace.id, site=site,
                        display_name=site, enabled=enabled,
                        hidden=False, sort_order=0,
                        target_sku_count=target_sku_count)
    db.add(row)
    db.flush()
    return row


def _product(db, site: str, sku: str) -> Product:
    row = Product(site=site, brand=site.split("_", 1)[0], sku=sku,
                  title=f"{sku} title", sale_price=10.0,
                  category_path="Storage", status="on_sale",
                  image_urls=["https://example.com/product.jpg"],
                  review_count=0)
    db.add(row)
    db.flush()
    return row


def _seed_two_workspaces(db):
    ws_a = _workspace(db, "Workspace A", "workspace-a")
    ws_b = _workspace(db, "Workspace B", "workspace-b")
    _site(db, "site_a", "A")
    _site(db, "site_b", "B")
    _workspace_site(db, ws_a, "site_a")
    _workspace_site(db, ws_b, "site_b")
    alice = _user(db, "alice", ws_a)
    bob = _user(db, "bob", ws_b)
    admin = _user(db, "admin", ws_a, role="admin",
                  global_role="super_admin")
    _product(db, "site_a", "A-1")
    _product(db, "site_b", "B-1")
    db.add(Promotion(site="site_b", sku="B-1", promotion_type="coupon",
                     promotion_name="B promo"))
    db.commit()
    return ws_a, ws_b, alice, bob, admin


def test_workspace_views_filter_global_warehouse_by_enabled_sites():
    from app.api.routes import list_products, list_sites

    db = _session()
    ws_a, ws_b, alice, bob, _admin = _seed_two_workspaces(db)

    alice_sites = list_sites(user="alice", x_workspace_id=str(ws_a.id), db=db)
    assert [s["site"] for s in alice_sites] == ["site_a"]
    alice_products = list_products(user="alice", x_workspace_id=str(ws_a.id),
                                   db=db)
    assert [p["sku"] for p in alice_products["items"]] == ["A-1"]

    bob_products = list_products(user="bob", x_workspace_id=str(ws_b.id), db=db)
    assert [p["sku"] for p in bob_products["items"]] == ["B-1"]

    with pytest.raises(HTTPException) as exc:
        list_products(site="site_b", user="alice",
                      x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 404
    assert db.query(Product).count() == 2
    assert not hasattr(Product, "workspace_id")


def test_product_trend_returns_same_listing_sku_variants():
    from app.api.routes import product_trend

    db = _session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    p1 = db.query(Product).filter(Product.site == "site_a",
                                  Product.sku == "A-1").one()
    p1.spu = "LISTING-A"
    p1.variant_id = "red"
    p2 = Product(site="site_a", brand="A", sku="A-2", spu="LISTING-A",
                 variant_id="blue", title="A-2 title", sale_price=12.0,
                 category_path="Storage", status="on_sale",
                 attributes={"color": "blue"})
    db.add(p2)
    db.add(PriceHistory(site="site_a", sku="A-1", date=date(2026, 1, 1),
                        sale_price=10.0, original_price=14.0, review_count=10))
    db.add(PriceHistory(site="site_a", sku="A-1", date=date(2026, 1, 2),
                        sale_price=9.0, original_price=14.0, review_count=12))
    db.add(Promotion(site="site_a", sku="A-1",
                     promotion_type="coupon",
                     promotion_name="Primary SKU coupon"))
    db.add(Promotion(site="site_a", sku="A-2",
                     promotion_type="price_promotion",
                     promotion_name="Variant SKU sale"))
    db.commit()

    payload = product_trend(pid=p1.id, user="alice", granularity="day",
                            x_workspace_id=str(ws_a.id), db=db)

    assert payload["product"]["sku"] == "A-1"
    assert [row["sku"] for row in payload["variants"]] == ["A-1", "A-2"]
    assert payload["variants"][1]["attributes"] == {"color": "blue"}
    assert payload["trend"][-1]["estimated_sales"] > 0
    assert {row["promotion_name"] for row in payload["promotions"]} == {
        "Primary SKU coupon", "Variant SKU sale",
    }

    sku_filtered = product_trend(pid=p1.id, user="alice", granularity="day",
                                 promo_sku="A-2",
                                 x_workspace_id=str(ws_a.id), db=db)
    assert [row["promotion_name"] for row in sku_filtered["promotions"]] == [
        "Variant SKU sale",
    ]
    outside_sku = product_trend(pid=p1.id, user="alice", granularity="day",
                                promo_sku="OTHER-SKU",
                                x_workspace_id=str(ws_a.id), db=db)
    assert outside_sku["promotions"] == []


def test_jobs_list_reports_total_and_summary_beyond_current_limit():
    from app.api.routes import list_jobs

    db = _threadsafe_session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    for idx in range(5):
        db.add(CrawlJob(site="site_a",
                        status="success" if idx < 4 else "running",
                        products_count=idx + 1,
                        requested_by_workspace_id=ws_a.id,
                        created_at=datetime(2026, 1, idx + 1),
                        finished_at=datetime(2026, 1, idx + 1, 1)
                        if idx < 4 else None))
    db.commit()

    payload = list_jobs(
        limit=2,
        created_from="2026-01-01T00:00:00+00:00",
        created_to="2026-01-05T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )

    assert payload["total"] == 5
    assert payload["page_size"] == 2
    assert len(payload["items"]) == 2
    assert payload["summary"]["success"] == 4
    assert payload["summary"]["running"] == 1
    assert payload["items"][0]["finished_at"] is None

    second_page = list_jobs(
        limit=2,
        page=2,
        created_from="2026-01-01T00:00:00+00:00",
        created_to="2026-01-05T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )
    assert second_page["page"] == 2
    assert second_page["page_size"] == 2
    assert [item["products_count"] for item in second_page["items"]] == [3, 2]


def test_jobs_list_collapses_same_site_same_day_by_default():
    from app.api.routes import list_jobs

    db = _threadsafe_session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    base = datetime(2026, 6, 25, 8, 0)
    previous_day = CrawlJob(site="site_a", status="success",
                            products_count=7,
                            requested_by_workspace_id=ws_a.id,
                            created_at=base - timedelta(days=1),
                            finished_at=base - timedelta(days=1) + timedelta(hours=1))
    old_same_day = CrawlJob(site="site_a", status="success",
                            products_count=10,
                            requested_by_workspace_id=ws_a.id,
                            created_at=base,
                            finished_at=base + timedelta(minutes=30))
    latest_same_day = CrawlJob(site="site_a", status="failed",
                               products_count=12,
                               requested_by_workspace_id=ws_a.id,
                               created_at=base + timedelta(hours=1),
                               finished_at=base + timedelta(hours=1, minutes=10))
    active_same_day = CrawlJob(site="site_a", status="running",
                               products_count=3,
                               requested_by_workspace_id=ws_a.id,
                               created_at=base + timedelta(hours=2),
                               started_at=base + timedelta(hours=2))
    db.add_all([previous_day, old_same_day, latest_same_day, active_same_day])
    db.commit()

    payload = list_jobs(
        limit=20,
        created_from="2026-06-24T00:00:00+00:00",
        created_to="2026-06-25T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )

    displayed_ids = {item["id"] for item in payload["items"]}
    assert payload["total"] == 2
    assert displayed_ids == {previous_day.id, active_same_day.id}
    assert payload["summary"]["success"] == 1
    assert payload["summary"]["running"] == 1
    assert payload["summary"]["active"] == 1

    exact_payload = list_jobs(
        limit=20,
        ids=f"{old_same_day.id},{latest_same_day.id}",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )
    assert exact_payload["total"] == 2
    assert {item["id"] for item in exact_payload["items"]} == {
        old_same_day.id,
        latest_same_day.id,
    }


def test_jobs_list_daily_collapse_prefers_full_site_job_over_failed_retry():
    from app.api.routes import list_jobs

    db = _threadsafe_session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    base = datetime(2026, 6, 25, 8, 0)
    full_site = CrawlJob(site="site_a", status="success", trigger="manual",
                         products_count=14428, total_product_count=14432,
                         requested_by_workspace_id=ws_a.id,
                         created_at=base,
                         finished_at=base + timedelta(minutes=30))
    failed_with_larger_denominator = CrawlJob(
        site="site_a", status="failed", trigger="manual",
        products_count=300, total_product_count=15000,
        failure_code="job_timeout",
        requested_by_workspace_id=ws_a.id,
        created_at=base + timedelta(minutes=40),
        finished_at=base + timedelta(minutes=45),
    )
    smaller_admin_retry = CrawlJob(site="site_a", status="success",
                                   trigger="admin_retry",
                                   products_count=1534,
                                   total_product_count=1535,
                                   requested_by_workspace_id=ws_a.id,
                                   created_at=base + timedelta(hours=1),
                                   finished_at=base + timedelta(hours=1, minutes=10))
    failed_retry = CrawlJob(site="site_a", status="success",
                            trigger="failed_product_retry",
                            products_count=500, total_product_count=500,
                            requested_by_workspace_id=ws_a.id,
                            created_at=base + timedelta(hours=2),
                            finished_at=base + timedelta(hours=2, minutes=5))
    db.add_all([
        full_site,
        failed_with_larger_denominator,
        smaller_admin_retry,
        failed_retry,
    ])
    db.commit()

    payload = list_jobs(
        limit=20,
        created_from="2026-06-25T00:00:00+00:00",
        created_to="2026-06-25T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )

    assert payload["total"] == 1
    assert payload["items"][0]["id"] == full_site.id
    assert payload["items"][0]["products_count"] == 14428
    assert payload["items"][0]["total_product_count"] == 14432


def test_jobs_list_date_range_collapses_by_default_but_can_show_history():
    from app.api.routes import list_jobs

    db = _threadsafe_session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    base = datetime(2026, 6, 25, 8, 0)
    first = CrawlJob(site="site_a", status="success",
                     requested_by_workspace_id=ws_a.id,
                     created_at=base,
                     finished_at=base + timedelta(minutes=30))
    second = CrawlJob(site="site_a", status="running",
                      requested_by_workspace_id=ws_a.id,
                      created_at=base + timedelta(hours=1),
                      started_at=base + timedelta(hours=1))
    outside = CrawlJob(site="site_a", status="success",
                       requested_by_workspace_id=ws_a.id,
                       created_at=base - timedelta(days=1),
                       finished_at=base - timedelta(days=1, minutes=-30))
    db.add_all([first, second, outside])
    db.commit()

    payload = list_jobs(
        limit=20,
        status="running",
        created_from="2026-06-25T00:00:00+00:00",
        created_to="2026-06-25T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )

    assert payload["total"] == 1
    assert [item["id"] for item in payload["items"]] == [second.id]
    assert payload["summary"]["running"] == 1
    assert payload["summary"].get("success", 0) == 0
    assert payload["total_all_statuses"] == 1
    assert payload["summary_all_statuses"]["running"] == 1
    assert payload["summary_all_statuses"].get("success", 0) == 0
    assert payload["site_scope"]["total"] == 1
    assert payload["site_scope"]["paused"] == 0
    assert payload["site_scope"]["trackable"] == 1

    history_payload = list_jobs(
        limit=20,
        created_from="2026-06-25T00:00:00+00:00",
        created_to="2026-06-25T23:59:59+00:00",
        collapse_daily=False,
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )
    assert history_payload["total"] == 2
    assert {item["id"] for item in history_payload["items"]} == {
        first.id,
        second.id,
    }
    assert history_payload["summary_all_statuses"]["running"] == 1
    assert history_payload["summary_all_statuses"]["success"] == 1

    db.query(Site).filter(Site.site == "site_b").one().track_status = "paused"
    db.commit()
    global_payload = list_jobs(
        limit=20,
        created_from="2026-06-25T00:00:00+00:00",
        created_to="2026-06-25T23:59:59+00:00",
        all_workspaces=True,
        user="admin",
        x_workspace_id=str(ws_a.id),
        db=db,
    )
    assert global_payload["global_view"] is True
    assert global_payload["site_scope"]["total"] == 2
    assert global_payload["site_scope"]["paused"] == 1
    assert global_payload["site_scope"]["trackable"] == 1


def test_jobs_list_hides_paused_tracking_sites():
    from app.api.routes import list_jobs

    db = _threadsafe_session()
    ws_a, ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    db.query(Site).filter(Site.site == "site_b").one().track_status = "paused"
    _site(db, "site_error", "Error Site")
    _workspace_site(db, ws_a, "site_error")
    db.query(Site).filter(Site.site == "site_error").one().track_status = "error"
    base = datetime(2026, 6, 25, 8, 0)
    visible = CrawlJob(site="site_a", status="success",
                       requested_by_workspace_id=ws_a.id,
                       created_at=base,
                       finished_at=base + timedelta(minutes=30))
    paused = CrawlJob(site="site_b", status="running",
                      requested_by_workspace_id=ws_b.id,
                      created_at=base + timedelta(hours=1),
                      started_at=base + timedelta(hours=1))
    hidden = CrawlJob(site="site_a", status="skipped",
                      requested_by_workspace_id=ws_a.id,
                      created_at=base + timedelta(hours=2),
                      finished_at=base + timedelta(hours=2, minutes=1),
                      failure_code="workspace_hidden")
    _site(db, "site_hidden", "Hidden")
    _workspace_site(db, ws_a, "site_hidden").hidden = True
    hidden_manual = CrawlJob(site="site_hidden", status="failed",
                             requested_by_workspace_id=ws_a.id,
                             created_at=base + timedelta(hours=3),
                             finished_at=base + timedelta(hours=3, minutes=1),
                             failure_code="zero_products")
    error_site = CrawlJob(site="site_error", status="failed",
                          requested_by_workspace_id=ws_a.id,
                          created_at=base + timedelta(hours=4),
                          finished_at=base + timedelta(hours=4, minutes=1),
                          failure_code="zero_products")
    db.add_all([visible, paused, hidden, hidden_manual, error_site])
    db.commit()

    payload = list_jobs(
        limit=20,
        created_from="2026-06-25T00:00:00+00:00",
        created_to="2026-06-25T23:59:59+00:00",
        all_workspaces=True,
        user="admin",
        x_workspace_id=str(ws_a.id),
        db=db,
    )

    assert payload["total_all_statuses"] == 2
    assert payload["summary_all_statuses"]["running"] == 0
    assert payload["summary_all_statuses"]["success"] == 1
    assert payload["summary_all_statuses"]["failed"] == 1
    assert [item["site"] for item in payload["items"]] == [
        "site_error", "site_a"
    ]
    assert payload["site_scope"]["total"] == 3
    assert payload["site_scope"]["paused"] == 1
    assert payload["site_scope"]["error"] == 1
    assert payload["site_scope"]["trackable"] == 1


def test_jobs_list_skips_live_progress_by_default():
    from app.api.routes import list_jobs

    db = _threadsafe_session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    created_at = datetime(2026, 6, 22, 10, 0)
    job = CrawlJob(site="site_a", status="running", products_count=2,
                   total_product_count=3,
                   requested_by_workspace_id=ws_a.id,
                   created_at=created_at)
    db.add(job)
    for idx in range(5):
        db.add(CrawlUrl(
            site="site_a",
            url=f"https://site-a.example.com/p/{idx}",
            url_hash=f"job-live-{idx}",
            kind="product",
            status="success" if idx < 4 else "pending",
            attempts=1 if idx < 4 else 0,
            last_seen_at=created_at + timedelta(minutes=idx),
        ))
    db.commit()

    payload = list_jobs(
        limit=20,
        created_from="2026-06-22T00:00:00+00:00",
        created_to="2026-06-22T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )
    row = payload["items"][0]
    assert row["products_count"] == 2
    assert row["total_product_count"] == 3
    assert row["total_product_count_source"] == "crawl_stats_total"

    job.products_count = 0
    job.total_product_count = 0
    db.commit()
    live_payload = list_jobs(
        limit=20,
        include_live_progress=True,
        created_from="2026-06-22T00:00:00+00:00",
        created_to="2026-06-22T23:59:59+00:00",
        user="alice",
        x_workspace_id=str(ws_a.id),
        db=db,
    )
    live_row = live_payload["items"][0]
    assert live_row["products_count"] == 4
    assert live_row["total_product_count"] == 5
    assert live_row["total_product_count_source"] == "crawl_frontier_live"


def test_members_can_switch_only_to_joined_workspaces():
    from app.api.routes import list_sites

    db = _session()
    ws_a, ws_b, alice, bob, _admin = _seed_two_workspaces(db)
    db.add(WorkspaceMember(workspace_id=ws_b.id, user_id=alice.id,
                           role="viewer", status="active"))
    db.commit()

    switched = list_sites(user="alice", x_workspace_id=str(ws_b.id), db=db)
    assert [s["site"] for s in switched] == ["site_b"]

    with pytest.raises(HTTPException) as exc:
        list_sites(user="bob", x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 403


def test_admin_users_respect_workspace_scope_for_super_admin():
    from app.api.routes import admin_list_users, admin_reset_password, admin_update_user

    db = _session()
    ws_a, ws_b, alice, bob, admin = _seed_two_workspaces(db)

    ws_a_users = admin_list_users(user="admin", x_workspace_id=str(ws_a.id), db=db)
    assert {u["username"] for u in ws_a_users} == {"alice", "admin"}
    assert all(u["workspace_ids"] == [ws_a.id] for u in ws_a_users)

    ws_b_users = admin_list_users(user="admin", x_workspace_id=str(ws_b.id), db=db)
    assert [u["username"] for u in ws_b_users] == ["bob"]
    assert ws_b_users[0]["workspace_ids"] == [ws_b.id]

    global_users = admin_list_users(user="admin", db=db)
    assert {u["username"] for u in global_users} == {"alice", "bob", "admin"}

    with pytest.raises(HTTPException) as exc:
        admin_update_user(bob.id, {"status": "disabled"},
                          user="admin", x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 404
    assert db.get(User, bob.id).status == "active"

    with pytest.raises(HTTPException) as exc:
        admin_reset_password(bob.id, {}, user="admin",
                             x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 404


def test_disabled_membership_cannot_fall_back_to_default_workspace():
    from app.api.routes import list_sites

    db = _session()
    ws_a, _ws_b, alice, _bob, _admin = _seed_two_workspaces(db)
    member = (db.query(WorkspaceMember)
              .filter(WorkspaceMember.workspace_id == ws_a.id,
                      WorkspaceMember.user_id == alice.id)
              .first())
    member.status = "disabled"
    db.commit()

    sites = list_sites(user="alice", db=db)
    assert sites == []


def test_tracking_write_permissions_use_workspace_role():
    from app.api.tracking import add_tracking

    db = _session()
    ws = _workspace(db, "Workspace A", "workspace-a")
    owner = _user(db, "owner_user", ws, role="user")
    admin_account = _user(db, "admin_account", ws, role="admin")
    owner_member = (db.query(WorkspaceMember)
                    .filter(WorkspaceMember.workspace_id == ws.id,
                            WorkspaceMember.user_id == owner.id)
                    .first())
    admin_member = (db.query(WorkspaceMember)
                    .filter(WorkspaceMember.workspace_id == ws.id,
                            WorkspaceMember.user_id == admin_account.id)
                    .first())
    owner_member.role = "owner"
    admin_member.role = "member"
    db.commit()

    with patch("app.api.tracking.detect_platform",
               return_value=("shopify", "https://newbrand.example.com")), \
         patch("app.api.tracking.enqueue", return_value=1):
        created = add_tracking(
            {"url": "https://newbrand.example.com/x", "brand": "NewBrand", "country": "US"},
            user="owner_user",
            x_workspace_id=str(ws.id),
            db=db,
        )
    assert created["source"] == "user"
    assert created["creator"] == "owner_user"

    with pytest.raises(HTTPException) as exc:
        add_tracking(
            {"url": "https://blocked.example.com"},
            user="admin_account",
            x_workspace_id=str(ws.id),
            db=db,
        )
    assert exc.value.status_code == 403


def test_report_config_and_exports_require_report_editor_role():
    from app.api.routes import create_report_config, export_promotions

    db = _session()
    ws = _workspace(db, "Workspace A", "workspace-a")
    _site(db, "site_a", "A")
    _workspace_site(db, ws, "site_a")
    viewer = _user(db, "report_viewer", ws, role="viewer")
    owner = _user(db, "report_owner", ws, role="user")
    owner_member = (db.query(WorkspaceMember)
                    .filter(WorkspaceMember.workspace_id == ws.id,
                            WorkspaceMember.user_id == owner.id)
                    .first())
    owner_member.role = "owner"
    db.add(Promotion(site="site_a", sku="SKU-1",
                     promotion_type="coupon",
                     promotion_name="Owner promo",
                     detected_time=datetime(2026, 6, 1)))
    db.commit()

    with pytest.raises(HTTPException) as exc:
        create_report_config({"name": "Blocked"},
                             user=viewer.username,
                             x_workspace_id=str(ws.id),
                             db=db)
    assert exc.value.status_code == 403

    created = create_report_config({"name": "Owner report",
                                    "sites": ["site_a"]},
                                   user=owner.username,
                                   x_workspace_id=str(ws.id),
                                   db=db)
    assert created["name"] == "Owner report"

    with pytest.raises(HTTPException) as exc:
        export_promotions(token=make_token(viewer.username, ""),
                          workspace_id=ws.id,
                          site="site_a",
                          db=db)
    assert exc.value.status_code == 403

    response = export_promotions(token=make_token(owner.username, ""),
                                 workspace_id=ws.id,
                                 site="site_a",
                                 db=db)
    assert response.status_code == 200


def test_super_admin_can_manage_cross_workspace_api_keys_and_usage_views():
    from app.api.routes import create_key, list_keys

    db = _session()
    ws_a, ws_b, _alice, _bob, admin = _seed_two_workspaces(db)

    created = create_key({"name": "workspace-b-key",
                          "workspace_id": ws_b.id,
                          "scopes": ["crawler:read"]},
                         user="admin", x_workspace_id=str(ws_a.id), db=db)

    assert created["workspace_id"] == ws_b.id
    assert [k["name"] for k in list_keys(user="admin",
                                         x_workspace_id=str(ws_a.id),
                                         db=db)] == []
    assert [k["name"] for k in list_keys(user="admin",
                                         x_workspace_id=str(ws_b.id),
                                         db=db)] == ["workspace-b-key"]

    stored = db.get(ApiKey, created["id"])
    stored.key_hash = hash_key("sck_workspace_b")
    db.commit()


def test_null_workspace_api_keys_are_not_listed_after_tenant_cutover():
    from app.api.routes import billing_usage, list_keys

    db = _session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    db.add(ApiKey(name="legacy-null", key_prefix="sck_null",
                  key_hash=hash_key("sck_null_secret"), active=True,
                  workspace_id=None))
    db.commit()

    assert list_keys(user="admin", x_workspace_id=str(ws_a.id), db=db) == []
    assert billing_usage(user="admin", x_workspace_id=str(ws_a.id),
                         db=db)["keys"] == []


def test_admin_cannot_assign_api_key_owner_outside_workspace():
    from app.api.routes import create_key, update_key

    db = _session()
    ws_a, ws_b, alice, bob, _admin = _seed_two_workspaces(db)

    with pytest.raises(HTTPException) as exc:
        create_key({"name": "bad-owner", "owner_user_id": bob.id},
                   user="admin", x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 400

    created = create_key({"name": "alice-key", "owner_user_id": alice.id},
                         user="admin", x_workspace_id=str(ws_a.id), db=db)
    with pytest.raises(HTTPException) as exc:
        update_key(created["id"], {"workspace_id": ws_b.id},
                   user="admin", x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 400


def test_invite_registration_joins_invite_workspace():
    from app.api.routes import auth_register

    db = _session()
    ws_a = _workspace(db, "Workspace A", "workspace-a")
    ws_b = _workspace(db, "Workspace B", "workspace-b")
    _user(db, "admin", ws_a, role="admin", global_role="super_admin")
    raw_code = "workspace-b-invite"
    db.add(InviteCode(code_prefix=raw_code[:10],
                      code_hash=hash_secret(raw_code),
                      workspace_id=ws_b.id,
                      max_uses=1,
                      used_count=0,
                      active=True,
                      default_role="user"))
    db.commit()

    registered = auth_register({
        "username": "carol",
        "email": "carol@example.com",
        "password": "Password1",
        "confirm_password": "Password1",
        "invite_code": raw_code,
    }, request=None, db=db)

    carol = db.query(User).filter(User.username == "carol").first()
    assert registered["username"] == "carol"
    assert carol.default_workspace_id == ws_b.id
    assert (db.query(WorkspaceMember)
            .filter(WorkspaceMember.workspace_id == ws_b.id,
                    WorkspaceMember.user_id == carol.id)
                    .count()) == 1


def test_token_export_preview_honors_requested_workspace_and_site_scoped_details():
    from app.api.routes import export_preview

    db = _session()
    ws_a, ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    db.add(PriceHistory(site="site_a", sku="SHARED", date=date.today(),
                        sale_price=10, original_price=12))
    db.add(PriceHistory(site="site_b", sku="SHARED", date=date.today(),
                        sale_price=20, original_price=24))
    db.add(Review(platform="trustpilot", site="site_a", sku="SHARED",
                  review_id="ra", review_date=datetime.utcnow(),
                  content="A"))
    db.add(Review(platform="trustpilot", site="site_b", sku="SHARED",
                  review_id="rb", review_date=datetime.utcnow(),
                  content="B"))
    _product(db, "site_a", "SHARED")
    _product(db, "site_b", "SHARED")
    db.commit()
    token = make_token(admin.username)

    preview_a = export_preview(token=token, workspace_id=ws_a.id,
                               include_price_history=True, include_voc=True,
                               db=db)
    preview_b = export_preview(token=token, workspace_id=ws_b.id,
                               include_price_history=True, include_voc=True,
                               db=db)

    assert preview_a["sku_count"] == 2
    assert preview_b["sku_count"] == 2
    assert preview_a["price_history_rows"] == 1
    assert preview_b["price_history_rows"] == 1
    assert preview_a["review_count"] == 1
    assert preview_b["review_count"] == 1

    with pytest.raises(HTTPException) as exc:
        export_preview(token=token, site="site_b", workspace_id=ws_a.id, db=db)
    assert exc.value.status_code == 404


def test_revoked_session_token_cannot_use_public_export_preview():
    from app.api.routes import auth_login, auth_logout, export_preview

    db = _session()
    ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    logged_in = auth_login({"identifier": "admin", "password": "Password1"},
                           request=None, db=db)

    auth_logout(authorization=f"Bearer {logged_in['token']}", db=db)

    with pytest.raises(HTTPException) as exc:
        export_preview(token=logged_in["token"], workspace_id=ws_a.id, db=db)
    assert exc.value.status_code == 401


def test_export_builders_do_not_leak_same_sku_across_sites():
    from app.export import (price_history_df, products_full_df,
                            reviews_voc_df, sites_overview_df)

    db = _session()
    _ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    _product(db, "site_a", "SHARED")
    _product(db, "site_b", "SHARED")
    _site(db, "export_de", "Export")
    export_product = _product(db, "export_de", "DE-MISSING-CURRENCY")
    export_product.currency = None
    db.add(PriceHistory(site="site_a", sku="SHARED", date=date.today(),
                        sale_price=10, original_price=12))
    db.add(PriceHistory(site="site_b", sku="SHARED", date=date.today(),
                        sale_price=20, original_price=24))
    db.add(Review(platform="trustpilot", site="site_a", sku="SHARED",
                  review_id="ra", review_date=datetime.utcnow(),
                  content="A"))
    db.add(Review(platform="trustpilot", site="site_b", sku="SHARED",
                  review_id="rb", review_date=datetime.utcnow(),
                  content="B"))
    db.commit()

    assert set(price_history_df(db, ["site_a"])["site"]) == {"site_a"}
    assert set(reviews_voc_df(db, ["site_a"])["site"]) == {"site_a"}
    assert set(sites_overview_df(db, ["site_a"])["site"]) == {"site_a"}
    full = products_full_df(db, ["export_de"])
    assert full.loc[full["sku"] == "DE-MISSING-CURRENCY", "currency"].iloc[0] == "EUR"


def test_site_overview_aggregates_trends_by_granularity_and_period():
    from app.api.routes import site_overview

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    db.add_all([
        Trend(site="site_a", date=date(2026, 5, 30), sku_count=10,
              new_product_count=1, estimated_sales=20,
              estimated_revenue=200, avg_rating=4.1, review_total=100),
        Trend(site="site_a", date=date(2026, 5, 31), sku_count=12,
              new_product_count=2, estimated_sales=30,
              estimated_revenue=300, avg_rating=4.2, review_total=120),
        Trend(site="site_a", date=date(2026, 6, 15), sku_count=20,
              new_product_count=3, estimated_sales=50,
              estimated_revenue=500, avg_rating=4.4, review_total=180,
              traffic=1234, conversion_rate=2.5),
    ])
    db.commit()

    monthly = site_overview("site_a", user=admin.username,
                            x_workspace_id=str(ws_a.id), db=db,
                            granularity="month")
    assert [row["date"] for row in monthly["trends"]] == ["2026-05", "2026-06"]
    assert monthly["trends"][0]["source_date"] == "2026-05-31"
    assert monthly["cards"]["traffic"] == 1234
    assert monthly["cards"]["conversion_rate"] == 2.5
    assert monthly["trend_summary"]["current_period"]["estimated_sales"] == 50
    assert monthly["trend_summary"]["current_period"]["traffic"] == 1234
    assert monthly["trend_summary"]["current_period"]["conversion_rate"] == 2.5
    assert monthly["trend_summary"]["previous_period"]["estimated_sales"] == 30

    filtered = site_overview("site_a", user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db,
                             granularity="week",
                             date_from="2026-06-01", date_to="2026-06-30")
    assert len(filtered["trends"]) == 1
    assert filtered["trends"][0]["date"].startswith("2026-W")
    assert filtered["trend_summary"]["visible_points"] == 1


def test_site_overview_uses_current_snapshot_when_trends_missing():
    from app.api.routes import site_overview

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    p = db.query(Product).filter(Product.site == "site_a",
                                 Product.sku == "A-1").one()
    p.thirty_day_sales = 7
    p.thirty_day_revenue = 88.5
    p.ratings = 4.6
    p.review_count = 12
    p.updated_time = datetime(2026, 6, 15, 10, 30)
    db.commit()

    out = site_overview("site_a", user=admin.username,
                        x_workspace_id=str(ws_a.id), db=db,
                        granularity="month")

    assert out["trends"] == [{
        "date": "2026-06",
        "source_date": "2026-06-15",
        "sku_count": 1,
        "product_count": 1,
        "spu_count": 1,
        "new_product_count": 1,
        "estimated_sales": 7,
        "estimated_revenue": 88.5,
        "traffic": None,
        "conversion_rate": None,
        "avg_rating": 4.6,
        "review_total": 12,
        "snapshot": True,
    }]
    assert out["trend_summary"]["snapshot_fallback"] is True
    assert out["trend_summary"]["current_period"]["estimated_sales"] == 7
    assert out["trend_summary"]["visible_points"] == 1


def test_site_overview_exposes_product_count_separate_from_sku_rows():
    from app.api.routes import site_overview

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    p1 = db.query(Product).filter(Product.site == "site_a",
                                  Product.sku == "A-1").one()
    p1.spu = "LISTING-1"
    p2 = _product(db, "site_a", "A-2")
    p2.spu = "LISTING-1"
    db.commit()

    out = site_overview("site_a", user=admin.username,
                        x_workspace_id=str(ws_a.id), db=db)

    assert out["cards"]["sku_count"] == 2
    assert out["cards"]["product_count"] == 1
    assert out["cards"]["spu_count"] == 1
    assert out["trend_summary"]["current_period"]["sku_count"] == 2
    assert out["trend_summary"]["current_period"]["product_count"] == 1


def test_site_sku_counts_are_consistent_across_report_endpoints():
    from app.api.routes import data_quality, list_products, list_sites, site_overview

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    p1 = db.query(Product).filter(Product.site == "site_a",
                                  Product.sku == "A-1").one()
    p1.spu = "LISTING-1"
    p2 = _product(db, "site_a", "A-2")
    p2.spu = "LISTING-1"
    p3 = _product(db, "site_a", "A-3")
    p3.spu = None
    db.add(Promotion(site="site_a", sku="A-1",
                     promotion_type="coupon",
                     promotion_name="Summer coupon",
                     detected_time=datetime(2026, 6, 16)))
    db.commit()

    sites = list_sites(user=admin.username, x_workspace_id=str(ws_a.id), db=db)
    site_row = next(row for row in sites if row["site"] == "site_a")
    overview = site_overview("site_a", user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    products = list_products(site="site_a", user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    quality = data_quality(user=admin.username, x_workspace_id=str(ws_a.id),
                           db=db)
    quality_row = next(row for row in quality["items"] if row["site"] == "site_a")

    assert site_row["sku_count"] == 3
    assert site_row["spu_count"] == 2
    assert products["total"] == 3
    by_sku = {row["sku"]: row for row in products["items"]}
    assert by_sku["A-1"]["promotion_labels"] == ["Summer coupon"]
    assert overview["cards"]["sku_count"] == 3
    assert overview["cards"]["product_count"] == 2
    assert overview["cards"]["spu_count"] == 2
    assert quality_row["sku_count"] == 3
    assert quality_row["spu_count"] == 2


def test_latest_products_tab_uses_available_timestamps_without_zeroing_totals():
    from app.api.routes import list_products, site_overview

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    existing = db.query(Product).filter(Product.site == "site_a",
                                        Product.sku == "A-1").one()
    existing.created_time = None
    existing.published_at = None
    existing.updated_time = datetime(2026, 6, 16, 10)
    older = _product(db, "site_a", "OLDER-ONLY")
    older.created_time = None
    older.published_at = None
    older.updated_time = datetime(2026, 5, 1, 10)
    no_time = _product(db, "site_a", "NO-TIME")
    no_time.created_time = None
    no_time.published_at = None
    no_time.updated_time = None
    db.commit()

    overview = site_overview("site_a", user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    latest = list_products(site="site_a", tab="new", user=admin.username,
                           x_workspace_id=str(ws_a.id), db=db)

    assert overview["cards"]["sku_count"] == 3
    assert overview["cards"]["latest_product_count"] == 2
    assert latest["total"] == 2
    assert {row["sku"] for row in latest["items"]} == {"A-1", "OLDER-ONLY"}


def test_site_overview_exposes_currency_and_real_update_time():
    from app.api.routes import _currency_for_site, list_products, site_overview

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "songmics_de", "Songmics")
    _workspace_site(db, ws_a, "songmics_de")
    product = _product(db, "songmics_de", "DE-1")
    product.thirty_day_revenue = 99.5
    product.currency = "USD"
    product.updated_time = datetime(2026, 6, 15, 10, 30)
    product.created_time = None
    product.published_at = None
    old_flagged = _product(db, "songmics_de", "DE-OLD")
    old_flagged.is_new = True
    old_flagged.created_time = datetime(2026, 6, 1)
    old_flagged.updated_time = datetime(2020, 1, 1)
    db.commit()

    out = site_overview("songmics_de", user=admin.username,
                        x_workspace_id=str(ws_a.id), db=db)
    assert out["currency"] == "EUR"
    assert out["cards"]["currency"] == "EUR"
    assert out["cards"]["new_product_count"] == 1
    assert out["cards"]["latest_product_count"] == 2
    assert out["updated_at"] == "2026-06-15T10:30:00"
    latest = list_products(site="songmics_de", tab="new", user=admin.username,
                           x_workspace_id=str(ws_a.id), db=db)
    assert latest["total"] == 2
    assert latest["items"][0]["sku"] == "DE-1"
    assert latest["items"][0]["currency"] == "EUR"
    assert _currency_for_site("mercadolibre_ar") == "ARS"
    assert _currency_for_site("lazada_my") == "MYR"
    assert _currency_for_site("shopee_vn") == "VND"


def test_workspace_site_target_sku_drives_coverage_and_quality_deviation():
    from app.api.routes import (admin_update_workspace_site, data_coverage,
                                data_quality)

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    ws_site = (db.query(WorkspaceSite)
               .filter(WorkspaceSite.workspace_id == ws_a.id,
                       WorkspaceSite.site == "site_a")
               .one())

    updated = admin_update_workspace_site(
        ws_a.id, ws_site.id, {"target_sku_count": 10},
        user=admin.username, db=db,
    )
    coverage = data_coverage(user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in coverage["sites"] if item["site"] == "site_a")
    quality = data_quality(user=admin.username,
                           x_workspace_id=str(ws_a.id), db=db)
    quality_row = next(item for item in quality["items"] if item["site"] == "site_a")

    assert updated["target_sku_count"] == 10
    assert row["target_sku_count"] == 10
    assert row["estimated_full"] == 10
    assert row["actual_product_count"] == 1
    assert row["actual_product_count_source"] == "product_listing"
    assert row["sku_deviation_pct"] == -90
    assert coverage["summary"]["high_deviation_count"] == 1
    assert quality_row["target_sku_count"] == 10
    assert quality_row["sku_deviation_pct"] == -90
    assert "sku_deviation_high" in quality_row["issues"]
    assert quality["summary"]["high_deviation"] == 1


def test_coverage_actual_product_count_can_use_discovered_urls_without_details():
    from app.api.routes import data_coverage

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "site_c", "C")
    _workspace_site(db, ws_a, "site_c", target_sku_count=99)
    db.add(CrawlUrl(site="site_c", url="https://site-c.example.com/p/1",
                    url_hash="c1", kind="product", source="sitemap",
                    status="pending"))
    db.add(CrawlUrl(site="site_c", url="https://site-c.example.com/p/2",
                    url_hash="c2", kind="product", source="sitemap",
                    status="pending"))
    db.commit()

    coverage = data_coverage(user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in coverage["sites"] if item["site"] == "site_c")

    assert row["target_sku_count"] == 99
    assert row["estimated_full"] == 99
    assert row["actual_product_count"] == 2
    assert row["actual_product_count_source"] == "discovered_url"
    assert row["product_detail_count"] == 0


def test_coverage_exposes_detail_count_consistency_issue():
    from app.api.routes import data_coverage

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "site_c", "C")
    _workspace_site(db, ws_a, "site_c")
    db.add(SiteMetric(
        site="site_c",
        sku_count=1,
        product_listing_count=1,
        fetched_count=0,
        discovered_product_url_count=0,
        price_signal_count=3,
        review_signal_count=0,
        sales_signal_count=0,
        revenue_signal_count=0,
    ))
    db.commit()

    coverage = data_coverage(user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in coverage["sites"] if item["site"] == "site_c")

    assert row["sku_count"] == 1
    assert row["spu_count"] == 1
    assert row["actual_product_count"] == 1
    assert row["actual_product_count_source"] == "product_listing"
    assert row["detail_sku_count"] == 3
    assert row["product_detail_count"] == 3
    assert row["report_product_count"] == 3
    assert row["count_consistency_status"] == "warning"
    assert row["count_consistency_issues"] == ["detail_sku_gt_sku_count"]


def test_coverage_does_not_treat_sitemap_only_rows_as_report_details():
    from app.api.routes import data_coverage

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "overstock_us", "Overstock")
    _workspace_site(db, ws_a, "overstock_us")
    db.add(SiteMetric(
        site="overstock_us",
        sku_count=1_000_417,
        product_listing_count=1_000_417,
        fetched_count=0,
        discovered_product_url_count=41,
        price_signal_count=0,
        review_signal_count=0,
        sales_signal_count=0,
        revenue_signal_count=0,
    ))
    db.add(CrawlJob(
        site="overstock_us",
        status="success",
        trigger="scheduled",
        products_count=1000,
        total_product_count=1000,
    ))
    db.commit()

    coverage = data_coverage(user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in coverage["sites"]
               if item["site"] == "overstock_us")

    assert row["actual_product_count"] == 1000
    assert row["actual_product_count_source"] == "latest_success_job"
    assert row["metadata_product_count"] == 1_000_417
    assert row["product_listing_count"] == 1_000_417
    assert row["product_detail_count"] == 0
    assert row["report_product_count"] == 0
    assert row["current_raw"] == 0
    assert row["status"] == "empty"


def test_coverage_has_no_implicit_static_target():
    from app.api.routes import data_coverage, data_quality

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "costway_us", "Costway")
    _workspace_site(db, ws_a, "costway_us")
    _product(db, "costway_us", "CW-1")
    _site(db, "vidaxl_es", "VidaXL")
    _workspace_site(db, ws_a, "vidaxl_es")
    _product(db, "vidaxl_es", "VX-ES-1")
    db.commit()

    coverage = data_coverage(user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in coverage["sites"] if item["site"] == "costway_us")
    quality = data_quality(user=admin.username,
                           x_workspace_id=str(ws_a.id), db=db)
    quality_row = next(item for item in quality["items"] if item["site"] == "costway_us")

    assert row["target_sku_count"] is None
    assert row["target_sku_source"] is None
    assert row["sku_deviation_pct"] is None
    assert quality_row["target_sku_count"] is None
    assert quality_row["target_sku_source"] is None
    assert "sku_deviation_high" not in quality_row["issues"]
    vidaxl_row = next(item for item in coverage["sites"]
                      if item["site"] == "vidaxl_es")
    vidaxl_quality = next(item for item in quality["items"]
                          if item["site"] == "vidaxl_es")
    assert vidaxl_row["target_sku_count"] is None
    assert vidaxl_row["target_sku_source"] is None
    assert vidaxl_quality["target_sku_count"] is None
    assert "sku_deviation_high" not in vidaxl_quality["issues"]


def test_vidaxl_live_sitemap_total_is_estimate_without_target(monkeypatch):
    from app.api import routes
    from app.api.routes import data_coverage, data_quality

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "vidaxl_es", "VidaXL")
    _workspace_site(db, ws_a, "vidaxl_es")
    _product(db, "vidaxl_es", "VX-ES-1")
    db.commit()
    monkeypatch.setattr(routes, "_load_sitemap_totals",
                        lambda: {"vidaxl_es": 339702})

    coverage = data_coverage(user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    quality = data_quality(user=admin.username,
                           x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in coverage["sites"] if item["site"] == "vidaxl_es")
    quality_row = next(item for item in quality["items"] if item["site"] == "vidaxl_es")

    assert row["estimated_full"] == 339702
    assert row["sitemap_product_count"] == 339702
    assert row["actual_product_count"] == 1
    assert row["actual_product_count_source"] == "product_listing"
    assert row["target_sku_count"] is None
    assert row["target_sku_source"] is None
    assert quality_row["estimated_full"] == 339702
    assert quality_row["target_sku_count"] is None


def test_product_filters_and_export_rows_share_same_scope():
    import asyncio
    import io
    from openpyxl import load_workbook
    from app.api.routes import (
        export_products,
        list_products,
        _filtered_products_query,
        _product_order_cols,
    )
    from app.export import products_sample_df_from_rows

    db = _session()
    ws_a, _ws_b, alice, _bob, admin = _seed_two_workspaces(db)
    matching = _product(db, "site_a", "MATCH-1")
    matching.spu = "PARENT-1"
    matching.title = "Matching storage cabinet"
    matching.sale_price = 20
    matching.original_price = 30
    matching.category_path = "Garage / Cabinet"
    matching.ratings = 4.8
    matching.review_count = 120
    matching.thirty_day_sales = 12
    matching.thirty_day_revenue = 240
    matching.has_video = True
    matching.has_free_shipping = True
    matching.created_time = datetime(2026, 6, 1)
    matching.published_at = datetime(2026, 5, 20)
    matching.updated_time = datetime(2026, 6, 3, 12)
    matching.product_url = "https://site-a.example.com/products/cabinet-main"
    matching.attributes = {"material": "walnut", "color": "white"}
    matching.tags = ["storage-special"]
    sibling = _product(db, "site_a", "MATCH-2")
    sibling.spu = "PARENT-1"
    sibling.title = "Matching storage cabinet variant"
    sibling.sale_price = 21
    sibling.original_price = 31
    sibling.category_path = "Garage / Cabinet"
    sibling.ratings = 4.9
    sibling.review_count = 150
    sibling.thirty_day_sales = 15
    sibling.thirty_day_revenue = 315
    sibling.has_video = True
    sibling.has_free_shipping = True
    sibling.created_time = datetime(2026, 6, 2)
    sibling.updated_time = datetime(2026, 6, 2, 12)
    sibling.product_url = "https://site-a.example.com/products/cabinet-variant"
    sibling.attributes = {"material": "oak"}
    updated_only = _product(db, "site_a", "UPDATED-ONLY")
    updated_only.title = "Updated only launch record"
    updated_only.created_time = None
    updated_only.published_at = None
    updated_only.updated_time = datetime(2026, 5, 18, 9)
    ignored = _product(db, "site_a", "MISS-1")
    ignored.title = "Outdoor bench"
    ignored.sale_price = 8
    ignored.category_path = "Garden"
    ignored.ratings = 3.2
    db.commit()

    q = _filtered_products_query(
        db, ["site_a"], site="site_a", search="cabinet",
        category="Garage", min_price=10, min_rating=4.5,
        min_reviews=100, min_sales=10, min_revenue=200,
        min_variants=2, max_variants=2,
        has_video=True, free_shipping=True, created_from="2026-05-01",
    )
    rows = q.order_by(*_product_order_cols("all")).all()
    df = products_sample_df_from_rows(rows)

    assert [p.sku for p in rows] == ["MATCH-1", "MATCH-2"]
    assert df["SKU"].tolist() == ["MATCH-1", "MATCH-2"]
    assert df["Variants"].tolist() == [2, 2]
    assert df["Sales Price"].tolist() == ["20", "21"]
    assert df["Price"].tolist() == ["30", "31"]
    assert "Sales" in df.columns
    assert "Revenues" in df.columns
    assert "30-Day Sales" not in df.columns
    assert "30-Days Sales" not in df.columns
    assert "Updated Time" in df.columns
    assert "Update Time" not in df.columns
    assert df["Free shipping"].tolist() == ["YES", "YES"]
    assert df["Created Time"].tolist() == [
        "2026-05-20 00:00",
        "2026-06-02 00:00",
    ]
    assert df["Updated Time"].tolist() == [
        "2026-06-03 12:00",
        "2026-06-02 12:00",
    ]
    localized_product = Product(
        site="site_ca", sku="LOCAL-CA", title="Localized",
        sale_price=20, original_price=30, currency="$",
    )
    localized = products_sample_df_from_rows([localized_product])
    assert localized["Sales Price"].tolist() == ["CAD 20"]
    assert localized["Price"].tolist() == ["CAD 30"]

    published_rows = _filtered_products_query(
        db, ["site_a"], site="site_a", search="cabinet",
        created_from="2026-05-01", created_to="2026-05-31",
    ).order_by(*_product_order_cols("all")).all()
    assert [p.sku for p in published_rows] == ["MATCH-1"]
    api_rows = list_products(site="site_a", search="cabinet",
                             created_from="2026-05-01", created_to="2026-05-31",
                             user=alice.username,
                             x_workspace_id=str(ws_a.id), db=db)
    assert api_rows["total"] == 1
    assert api_rows["items"][0]["sku"] == "MATCH-1"
    assert api_rows["items"][0]["published_at"] == "2026-05-20T00:00:00"
    updated_only_rows = list_products(
        site="site_a", search="Updated only",
        created_from="2026-05-01", created_to="2026-05-31",
        user=alice.username, x_workspace_id=str(ws_a.id), db=db)
    assert updated_only_rows["total"] == 1
    assert updated_only_rows["items"][0]["sku"] == "UPDATED-ONLY"
    assert updated_only_rows["items"][0]["updated_time"] == "2026-05-18T09:00:00"

    url_rows = _filtered_products_query(
        db, ["site_a"], site="site_a", search="cabinet-main",
    ).order_by(*_product_order_cols("all")).all()
    attr_rows = _filtered_products_query(
        db, ["site_a"], site="site_a", search="walnut",
    ).order_by(*_product_order_cols("all")).all()
    tag_rows = _filtered_products_query(
        db, ["site_a"], site="site_a", search="storage-special",
    ).order_by(*_product_order_cols("all")).all()
    assert [p.sku for p in url_rows] == ["MATCH-1"]
    assert [p.sku for p in attr_rows] == ["MATCH-1"]
    assert [p.sku for p in tag_rows] == ["MATCH-1"]

    original_only = _product(db, "site_a", "ORIG-ONLY")
    original_only.title = "Original price only"
    original_only.sale_price = None
    original_only.original_price = 99
    db.commit()
    price_rows = _filtered_products_query(
        db, ["site_a"], site="site_a", search="Original price",
        min_price=90, max_price=100,
    ).order_by(*_product_order_cols("all")).all()
    assert [p.sku for p in price_rows] == ["ORIG-ONLY"]

    legacy_status = _product(db, "site_a", "STATUS-LEGACY")
    legacy_status.status = "out of stock"
    db.commit()
    status_rows = _filtered_products_query(
        db, ["site_a"], site="site_a", status="out_of_stock",
    ).order_by(*_product_order_cols("all")).all()
    assert [p.sku for p in status_rows] == ["STATUS-LEGACY"]

    response = export_products(
        token=make_token(admin.username, ""),
        workspace_id=ws_a.id,
        site="site_a",
        scope="products",
        export_scope="page",
        page=1,
        page_size=1,
        search="cabinet",
        category="Garage",
        min_price=10,
        min_rating=4.5,
        min_reviews=100,
        min_sales=10,
        min_revenue=200,
        min_variants=2,
        max_variants=2,
        has_video=True,
        free_shipping=True,
        created_from="2026-05-01",
        db=db,
    )

    async def _read_body(iterator):
        chunks = []
        async for chunk in iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    content = asyncio.run(_read_body(response.body_iterator))
    wb = load_workbook(io.BytesIO(content), read_only=True)
    headers = [cell.value for cell in next(wb["产品分析"].iter_rows(max_row=1))]
    assert headers[3] == "Products Details"
    exported = list(wb["产品分析"].iter_rows(min_row=2, values_only=True))
    assert [row[1] for row in exported] == ["MATCH-1"]
    assert exported[0][7] == 2


def test_promotion_filters_and_export_rows_share_same_scope():
    import asyncio
    import io
    from openpyxl import load_workbook
    from app.api.routes import (export_promotions, list_promotions,
                                _filtered_promotions_query)
    from app.export import promotions_sample_df_from_rows

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    db.add(Promotion(site="site_a", sku="PROMO-1",
                     promotion_type="coupon",
                     promotion_name="Cabinet coupon",
                     product_title="Storage Cabinet",
                     original_price=40,
                     promotion_price=30,
                     detected_time=datetime(2026, 6, 2)))
    db.add(Promotion(site="site_a", sku="PROMO-2",
                     promotion_type="bundle",
                     promotion_name="Bench bundle",
                     product_title="Outdoor Bench",
                     detected_time=datetime(2026, 6, 3)))
    db.add(Promotion(site="site_a", sku="PROMO-3",
                     promotion_type="price_promotion",
                     promotion_name="Flash price",
                     detected_time=datetime(2026, 6, 4)))
    promo_product_3 = _product(db, "site_a", "PROMO-3")
    promo_product_3.title = "Fallback Chair"
    promo_product_3.image_urls = ["https://site-a.example.com/chair.jpg"]
    promo_product = _product(db, "site_a", "PROMO-2")
    promo_product.product_url = "https://site-a.example.com/products/bench-url-token"
    db.commit()

    q = _filtered_promotions_query(
        db, ["site_a"], site="site_a", search="cabinet",
        type="coupon", date_from="2026-06-01", date_to="2026-06-30",
    )
    rows = q.order_by(Promotion.detected_time.desc().nullslast(),
                      Promotion.id.desc()).all()
    df = promotions_sample_df_from_rows(rows)

    assert [p.sku for p in rows] == ["PROMO-1"]
    assert df["SKU"].tolist() == ["PROMO-1"]
    assert df["Type"].tolist() == ["Coupons"]
    assert "Updated Time" in df.columns
    assert "Update Time" not in df.columns
    assert df["Updated Time"].tolist() == ["2026-06-02 00:00"]
    assert df["Discount"].tolist() == ["10"]
    assert df["Pre-price"].tolist() == ["40"]
    assert df["Post-price"].tolist() == ["30"]
    assert "Pre-price" in df.columns
    assert "Orignal-Price" not in df.columns

    localized = promotions_sample_df_from_rows([
        Promotion(site="promo_ca", sku="PROMO-CA", original_price=40,
                  promotion_price=30, detected_time=datetime(2026, 6, 2))
    ])
    assert localized["Discount"].tolist() == ["CAD 10"]
    assert localized["Pre-price"].tolist() == ["CAD 40"]
    assert localized["Post-price"].tolist() == ["CAD 30"]

    url_q = _filtered_promotions_query(
        db, ["site_a"], site="site_a", search="bench-url-token",
        type="bundle", date_from="2026-06-01", date_to="2026-06-30",
    )
    url_rows = url_q.order_by(Promotion.detected_time.desc().nullslast(),
                              Promotion.id.desc()).all()
    assert [p.sku for p in url_rows] == ["PROMO-2"]

    price_q = _filtered_promotions_query(
        db, ["site_a"], site="site_a", type="价格促销",
        date_from="2026-06-01", date_to="2026-06-30",
    )
    price_rows = price_q.order_by(Promotion.detected_time.desc().nullslast(),
                                  Promotion.id.desc()).all()
    assert [p.sku for p in price_rows] == ["PROMO-3"]

    page_one = list_promotions(site="site_a", date_from="2026-06-01",
                               date_to="2026-06-30", page=1, page_size=1,
                               user=admin.username,
                               x_workspace_id=str(ws_a.id), db=db)
    page_two = list_promotions(site="site_a", date_from="2026-06-01",
                               date_to="2026-06-30", page=2, page_size=1,
                               user=admin.username,
                               x_workspace_id=str(ws_a.id), db=db)
    assert page_one["total"] == 3
    assert page_one["page"] == 1
    assert page_one["page_size"] == 1
    assert [row["sku"] for row in page_one["items"]] == ["PROMO-3"]
    assert [row["sku"] for row in page_two["items"]] == ["PROMO-2"]

    response = export_promotions(
        token=make_token(admin.username, ""),
        workspace_id=ws_a.id,
        site="site_a",
        date_from="2026-06-01",
        date_to="2026-06-30",
        export_scope="page",
        page=1,
        page_size=1,
        db=db,
    )

    async def _read_body(iterator):
        chunks = []
        async for chunk in iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    content = asyncio.run(_read_body(response.body_iterator))
    wb = load_workbook(io.BytesIO(content), read_only=True)
    headers = [cell.value for cell in next(wb["销售促销"].iter_rows(max_row=1))]
    assert headers[:5] == ["NO.", "SKU", "Updated Time", "Products Details", "Product Image"]
    exported = list(wb["销售促销"].iter_rows(min_row=2, values_only=True))
    assert [row[1] for row in exported] == ["PROMO-3"]
    assert exported[0][3] == "Fallback Chair"
    assert exported[0][4] == "https://site-a.example.com/chair.jpg"


def test_workbook_export_backfills_promo_product_fields_and_trend_snapshot():
    import io
    from openpyxl import load_workbook
    from app.export import export_workbook

    db = _session()
    _ws_a, _ws_b, _alice, _bob, _admin = _seed_two_workspaces(db)
    product = db.query(Product).filter(Product.site == "site_a",
                                       Product.sku == "A-1").one()
    product.title = "Workbook Fallback Cabinet"
    product.image_urls = ["https://site-a.example.com/a-1.jpg"]
    product.thirty_day_sales = 7
    product.thirty_day_revenue = 140
    product.created_time = datetime(2026, 6, 1, 9, 0)
    product.updated_time = datetime(2026, 6, 5, 10, 30)
    db.add(Promotion(site="site_a", sku="A-1",
                     promotion_type="price_promotion",
                     original_price=30,
                     promotion_price=20,
                     detected_time=datetime(2026, 6, 6, 8, 0)))
    db.commit()

    content = export_workbook(db, site="site_a")
    wb = load_workbook(io.BytesIO(content), read_only=True)
    product_headers = [cell.value for cell in next(wb["商品分析"].iter_rows(max_row=1))]
    assert product_headers[3] == "Products Details"
    promo_headers = [cell.value for cell in next(wb["销售促销"].iter_rows(max_row=1))]
    assert promo_headers[:5] == ["NO.", "SKU", "Updated Time", "Products Details", "Product Image"]
    promo_rows = list(wb["销售促销"].iter_rows(min_row=2, values_only=True))
    trend_rows = list(wb["趋势报告"].iter_rows(min_row=2, values_only=True))

    assert promo_rows[0][1] == "A-1"
    assert promo_rows[0][3] == "Workbook Fallback Cabinet"
    assert promo_rows[0][4] == "https://site-a.example.com/a-1.jpg"
    assert trend_rows == [(1, "2026-06-05", 1, 1, 7, 140, "/", "/")]


def test_promotions_api_returns_product_url_for_clickable_sku():
    from app.api.routes import list_promotions

    db = _session()
    ws_a, _ws_b, alice, _bob, _admin = _seed_two_workspaces(db)
    product = db.query(Product).filter(Product.site == "site_a",
                                       Product.sku == "A-1").one()
    product.product_url = "https://site-a.example.com/products/a-1"
    product.title = "Linked Product"
    product.image_urls = ["https://site-a.example.com/a-1.jpg"]
    product.spu = "LISTING-A"
    product.label = "Storage"
    product.is_new = True
    product.is_bestseller = True
    db.add(Product(site="site_a", brand="A", sku="A-2", spu="LISTING-A",
                   variant_id="blue", title="Linked Product Blue",
                   product_url="https://site-a.example.com/products/a-2"))
    db.add(Promotion(site="site_a", sku="A-1",
                     promotion_type="price",
                     promotion_name="A sale",
                     detected_time=datetime(2026, 6, 4)))
    db.commit()

    out = list_promotions(site="site_a", user=alice.username,
                          x_workspace_id=str(ws_a.id), db=db)

    assert out["total"] == 1
    row = out["items"][0]
    assert row["product_url"] == "https://site-a.example.com/products/a-1"
    assert row["product_title"] == "Linked Product"
    assert row["product_image"] == "https://site-a.example.com/a-1.jpg"
    assert row["product_label"] == "Storage"
    assert row["is_new"] is True
    assert row["is_bestseller"] is True
    assert row["listing_sku"] == "A-1"
    assert row["variant_skus"] == ["A-1", "A-2"]
    assert row["variant_count"] == 2


def test_data_quality_surfaces_site_level_quality_gaps():
    from app.api.routes import data_quality

    db = _session()
    ws_a, ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    p = db.query(Product).filter(Product.site == "site_a", Product.sku == "A-1").first()
    p.sale_price = None
    p.original_price = None
    p.thirty_day_sales = 0
    p.thirty_day_revenue = 0
    p.title = p.sku
    db.add(CrawlJob(site="site_a", status="failed",
                    requested_by_workspace_id=ws_a.id,
                    failure_code="parse_none",
                    failure_stage="parse",
                    suggested_action="检查解析规则",
                    created_at=datetime(2026, 6, 1),
                    finished_at=datetime(2026, 6, 1, 0, 1)))
    db.commit()

    out = data_quality(user=admin.username, x_workspace_id=str(ws_a.id), db=db)
    rows = {row["site"]: row for row in out["items"]}

    assert set(rows) == {"site_a"}
    assert rows["site_a"]["status"] == "warning"
    assert "price_missing" in rows["site_a"]["issues"]
    assert "title_weak" in rows["site_a"]["issues"]
    assert "sales_missing" in rows["site_a"]["issues"]
    assert "revenue_missing" in rows["site_a"]["issues"]
    assert "traffic_missing" in rows["site_a"]["issues"]
    assert "conversion_missing" in rows["site_a"]["issues"]
    assert "promotions_missing" in rows["site_a"]["issues"]
    assert "latest_job_failed" in rows["site_a"]["issues"]
    assert rows["site_a"]["latest_job"]["failure_code"] == "parse_none"
    assert rows["site_a"]["last_error_code"] == "parse_none"
    assert rows["site_a"]["latest_failure"]["code"] == "parse_none"
    assert rows["site_a"]["severity"] == "warning"
    assert rows["site_a"]["suggestion"] == rows["site_a"]["suggested_action"]
    assert rows["site_a"]["external_data_required"] is True
    assert rows["site_a"]["rerun_blocked"] is True
    assert rows["site_a"]["rerun_recommended"] is False
    assert rows["site_a"]["rerun_ready"] is False
    assert rows["site_a"]["rerun_after_setup"] is True
    assert set(rows["site_a"]["rerun_preconditions"]) >= {
        "traffic_missing", "conversion_missing",
    }
    assert out["summary"]["needs_rerun"] == 0
    assert out["summary"]["rerun_after_setup"] == 1
    assert out["summary"]["external_data_required"] == 1
    assert out["summary"]["rerun_blocked"] == 1
    assert out["summary"]["rerun_precondition_total"] == 1
    preconditions = {
        item["issue"]: item for item in out["summary"]["rerun_preconditions"]
    }
    assert preconditions["traffic_missing"]["count"] == 1
    assert preconditions["traffic_missing"]["sites"] == ["site_a"]
    assert preconditions["conversion_missing"]["count"] == 1
    assert preconditions["conversion_missing"]["sites"] == ["site_a"]
    assert out["summary"]["missing_prices"] == 1
    assert out["summary"]["weak_titles"] == 1
    assert out["summary"]["missing_sales"] == 1
    assert out["summary"]["missing_traffic"] == 1
    assert out["summary"]["missing_conversion"] == 1
    assert rows["site_a"]["price_signal_count"] == 0
    assert rows["site_a"]["price_signal_pct"] == 0
    assert rows["site_a"]["weak_title_count"] == 1
    assert rows["site_a"]["title_quality_pct"] == 0
    assert rows["site_a"]["crawl_queue"]["failed"] == 1
    assert out["summary"]["failed_jobs"] == 1
    assert "价格解析" in rows["site_a"]["suggested_action"]

    p.sale_price = 10
    p.original_price = 12
    p.thirty_day_sales = 1
    p.thirty_day_revenue = 10
    p.title = "Real Product Title"
    promo = Promotion(site="site_a", sku="A-1", product_title="A-1 promo",
                      promotion_type="coupon", detected_time=datetime(2026, 6, 2))
    db.add(promo)
    db.add(Trend(site="site_a", date=date(2026, 6, 2),
                 sku_count=1, new_product_count=0,
                 estimated_sales=1, estimated_revenue=10,
                 traffic=100, conversion_rate=2.5))
    db.commit()
    out_ok_data_failed_job = data_quality(user=admin.username,
                                          x_workspace_id=str(ws_a.id), db=db)
    row = out_ok_data_failed_job["items"][0]
    assert row["status"] == "warning"
    assert row["issues"] == ["latest_job_failed"]
    assert row["crawl_queue"]["failed"] == 1
    assert "最近任务失败" in row["suggested_action"]

    db.add(CrawlJob(site="site_a", status="success",
                    requested_by_workspace_id=ws_a.id,
                    failure_code="http_429",
                    failure_stage="fetch",
                    failure_detail="HTTP 429 rate limited",
                    suggested_action="降低并发并切换住宅代理后重跑",
                    products_count=1,
                    created_at=datetime(2026, 6, 3),
                    finished_at=datetime(2026, 6, 3, 0, 1)))
    db.commit()
    partial_quality = data_quality(user=admin.username,
                                   x_workspace_id=str(ws_a.id), db=db)
    partial_row = partial_quality["items"][0]
    assert partial_row["latest_job"]["status"] == "success"
    assert partial_row["latest_job"]["failure_code"] == "http_429"
    assert "partial_crawl" in partial_row["issues"]
    assert "anti_bot_blocked" in partial_row["issues"]
    assert partial_quality["summary"]["partial_crawls"] == 1
    assert "住宅代理" in partial_row["suggested_action"]

    db.add(CrawlJob(site="site_a", status="blocked",
                    requested_by_workspace_id=ws_a.id,
                    failure_code="http_403",
                    failure_stage="fetch",
                    failure_detail="historical block",
                    suggested_action="切换住宅代理",
                    products_count=0,
                    created_at=datetime(2026, 6, 4),
                    finished_at=datetime(2026, 6, 4, 0, 1)))
    db.add(CrawlJob(site="site_a", status="success",
                    requested_by_workspace_id=ws_a.id,
                    products_count=1,
                    promotion_count=1,
                    created_at=datetime(2026, 6, 5),
                    finished_at=datetime(2026, 6, 5, 0, 1)))
    db.commit()
    recovered_quality = data_quality(user=admin.username,
                                     x_workspace_id=str(ws_a.id), db=db)
    recovered_row = recovered_quality["items"][0]
    assert recovered_row["status"] == "healthy"
    assert recovered_row["issues"] == []
    assert recovered_row["latest_failure"] is None
    assert recovered_row["crawl_queue"]["failed"] == 1
    assert recovered_row["crawl_queue"]["blocked"] == 1
    assert recovered_quality["summary"]["sites_with_failed_jobs"] == 1

    other = data_quality(user=admin.username, x_workspace_id=str(ws_b.id), db=db)
    assert {row["site"] for row in other["items"]} == {"site_b"}


def test_admin_data_quality_is_global_and_super_admin_only():
    from app.api.admin_spine import admin_data_quality

    db = _session()
    ws_a, ws_b, alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "site_empty", "E")
    _workspace_site(db, ws_a, "site_empty")
    db.add(CrawlJob(site="site_empty", status="failed",
                    requested_by_workspace_id=ws_a.id,
                    failure_code="zero_products",
                    failure_stage="parse",
                    created_at=datetime(2026, 6, 1),
                    finished_at=datetime(2026, 6, 1, 0, 1)))
    db.add(CrawlJob(site="site_a", status="pending",
                    requested_by_workspace_id=ws_a.id,
                    created_at=datetime.utcnow() - timedelta(hours=3)))
    db.add(CrawlJob(site="site_a", status="running",
                    requested_by_workspace_id=ws_a.id,
                    created_at=datetime.utcnow() - timedelta(hours=2),
                    started_at=datetime.utcnow() - timedelta(hours=2)))
    db.commit()

    out = admin_data_quality(user=admin.username, db=db)
    assert {row["site"] for row in out["items"]} == {"site_a", "site_b", "site_empty"}
    assert out["summary"]["workspace_count"] == 2
    site_a = next(row for row in out["items"] if row["site"] == "site_a")
    assert site_a["workspaces"] == [{"id": ws_a.id, "name": ws_a.name}]
    assert site_a["crawl_queue"]["pending"] == 1
    assert site_a["crawl_queue"]["running"] == 0
    assert site_a["crawl_queue"]["stuck"] == 1
    assert site_a["crawl_queue"]["active_count"] == 2
    assert site_a["crawl_queue"]["stale_pending"] == 1
    assert "job_pending_stale" in site_a["issues"]
    assert "排队超过久排阈值" in site_a["suggested_action"]
    assert out["summary"]["pending_jobs"] == 1
    assert out["summary"]["stuck_jobs"] == 1
    assert out["summary"]["stale_pending_jobs"] == 1
    site_empty = next(row for row in out["items"] if row["site"] == "site_empty")
    assert site_empty["status"] == "critical"
    assert site_empty["rerun_recommended"] is True
    assert out["summary"]["needs_rerun"] >= 1

    tenant_only = admin_data_quality(tenant=ws_b.id, user=admin.username, db=db)
    assert {row["site"] for row in tenant_only["items"]} == {"site_b"}
    assert tenant_only["summary"]["tenant_id"] == ws_b.id

    with pytest.raises(HTTPException) as exc:
        admin_data_quality(user=alice.username, db=db)
    assert exc.value.status_code == 403


def test_admin_queue_stats_list_and_detail_match_real_job_states():
    from app.api.admin_spine import job_detail, jobs_list, jobs_stats

    db = _session()
    ws_a, _ws_b, alice, _bob, admin = _seed_two_workspaces(db)
    now = datetime.utcnow()
    active = CrawlJob(site="site_a", status="running",
                      requested_by_workspace_id=ws_a.id,
                      trigger="manual",
                      products_count=7, new_count=2,
                      promotion_count=1,
                      worker="crawl-active",
                      created_at=now - timedelta(minutes=3),
                      started_at=now - timedelta(minutes=2))
    stuck = CrawlJob(site="site_a", status="running",
                     requested_by_workspace_id=ws_a.id,
                     trigger="scheduled",
                     worker="crawl-stuck",
                     created_at=now - timedelta(hours=2),
                     started_at=now - timedelta(hours=2))
    stale_pending = CrawlJob(site="site_a", status="pending",
                             requested_by_workspace_id=ws_a.id,
                             trigger="admin_quality_rerun",
                             created_at=now - timedelta(hours=3))
    failed = CrawlJob(site="site_a", status="failed",
                      requested_by_workspace_id=ws_a.id,
                      failure_code="parse_none",
                      failure_stage="parse",
                      failure_detail="no products parsed",
                      suggested_action="检查解析规则",
                      retryable=True,
                      created_at=now - timedelta(minutes=5),
                      finished_at=now - timedelta(minutes=4))
    spine_running = SpineJob(url="https://example.com/a",
                             dataset="catalog-feed",
                             entity_type="product",
                             status="running",
                             worker="spine-active",
                             workspace_id=ws_a.id,
                             created_at=now - timedelta(minutes=4),
                             started_at=now - timedelta(minutes=3),
                             heartbeat_at=now)
    spine_stuck = SpineJob(url="https://example.com/b",
                           dataset="catalog-feed",
                           entity_type="product",
                           status="running",
                           worker="spine-stuck",
                           workspace_id=ws_a.id,
                           created_at=now - timedelta(hours=2),
                           started_at=now - timedelta(hours=2),
                           heartbeat_at=now - timedelta(hours=2))
    spine_failed = SpineJob(url="https://example.com/c",
                            dataset="catalog-feed",
                            entity_type="product",
                            status="failed",
                            error="fetch timeout",
                            retries=2,
                            max_retries=3,
                            workspace_id=ws_a.id,
                            created_at=now - timedelta(minutes=6),
                            finished_at=now - timedelta(minutes=5))
    ondemand_running = OnDemandJob(url="https://shop.example/a",
                                   platform="lazada",
                                   kind="product",
                                   status="running",
                                   workspace_id=ws_a.id,
                                   batch_id="batch-live",
                                   attempts=1,
                                   created_at=now - timedelta(minutes=3))
    ondemand_stuck = OnDemandJob(url="https://shop.example/b",
                                 platform="lazada",
                                 kind="product",
                                 status="running",
                                 workspace_id=ws_a.id,
                                 batch_id="batch-old",
                                 attempts=2,
                                 created_at=now - timedelta(hours=2))
    ondemand_failed = OnDemandJob(url="https://shop.example/c",
                                  platform="shopee",
                                  kind="product",
                                  status="failed",
                                  error="captcha",
                                  workspace_id=ws_a.id,
                                  batch_id="batch-failed",
                                  attempts=3,
                                  created_at=now - timedelta(minutes=7),
                                  finished_at=now - timedelta(minutes=6))
    db.add_all([
        active, stuck, stale_pending, failed,
        spine_running, spine_stuck, spine_failed,
        ondemand_running, ondemand_stuck, ondemand_failed,
    ])
    db.commit()

    stats = jobs_stats(user=admin.username, db=db)
    assert stats["by_queue"]["crawl"]["running"] == 1
    assert stats["by_queue"]["crawl"]["stuck"] == 1
    assert stats["by_queue"]["crawl"]["stale_pending"] == 1
    assert stats["by_queue"]["crawl"]["failed"] == 1
    assert stats["by_queue"]["spine"]["running"] == 1
    assert stats["by_queue"]["spine"]["stuck"] == 1
    assert stats["by_queue"]["spine"]["failed"] == 1
    assert stats["by_queue"]["ondemand"]["running"] == 1
    assert stats["by_queue"]["ondemand"]["stuck"] == 1
    assert stats["by_queue"]["ondemand"]["failed"] == 1
    assert stats["breakdowns"]["crawl_running_by_site"] == [
        {"key": "site_a", "count": 1}
    ]
    assert stats["breakdowns"]["crawl_stuck_by_site"] == [
        {"key": "site_a", "count": 1}
    ]
    assert stats["breakdowns"]["crawl_stale_pending_by_site"] == [
        {"key": "site_a", "count": 1}
    ]
    assert stats["breakdowns"]["crawl_failed_by_site"] == [
        {"key": "site_a", "count": 1}
    ]
    assert stats["breakdowns"]["spine_running_by_dataset"] == [
        {"key": "catalog-feed", "count": 1}
    ]
    assert stats["breakdowns"]["spine_stuck_by_dataset"] == [
        {"key": "catalog-feed", "count": 1}
    ]
    assert stats["breakdowns"]["spine_failed_by_dataset"] == [
        {"key": "catalog-feed", "count": 1}
    ]
    assert stats["breakdowns"]["ondemand_running_by_platform"] == [
        {"key": "lazada", "count": 1}
    ]
    assert stats["breakdowns"]["ondemand_stuck_by_platform"] == [
        {"key": "lazada", "count": 1}
    ]
    assert stats["breakdowns"]["ondemand_failed_by_platform"] == [
        {"key": "shopee", "count": 1}
    ]

    running = jobs_list(status="running", dataset="site_a", source="crawl",
                        user=admin.username, db=db)
    assert running["total"] == 1
    assert running["items"][0]["id"] == active.id
    assert running["items"][0]["normalized_status"] == "running"

    stuck_rows = jobs_list(status="stuck", dataset="site_a", source="crawl",
                           user=admin.username, db=db)
    assert stuck_rows["total"] == 1
    assert stuck_rows["items"][0]["id"] == stuck.id
    assert stuck_rows["items"][0]["normalized_status"] == "stuck"
    assert stuck_rows["items"][0]["stuck_reason"] == "running_timeout"

    stale_rows = jobs_list(status="stale_pending", dataset="site_a",
                           source="crawl", user=admin.username, db=db)
    assert stale_rows["total"] == 1
    assert stale_rows["items"][0]["id"] == stale_pending.id
    assert stale_rows["items"][0]["is_stale_pending"] is True
    assert stale_rows["items"][0]["stuck_reason"] == "pending_too_long"

    failed_rows = jobs_list(status="failed", dataset="site_a",
                            failure_code="parse_none", source="crawl",
                            user=admin.username, db=db)
    assert failed_rows["total"] == 1
    assert failed_rows["items"][0]["id"] == failed.id
    assert failed_rows["items"][0]["failure_detail"] == "no products parsed"
    detail = job_detail(failed.id, source="crawl", user=admin.username, db=db)
    assert detail["failure_code"] == "parse_none"
    assert detail["suggested_action"] == "检查解析规则"
    assert detail["retryable"] is True

    spine_running_rows = jobs_list(status="running", dataset="catalog-feed",
                                   source="spine", user=admin.username, db=db)
    assert spine_running_rows["total"] == 1
    assert spine_running_rows["items"][0]["id"] == spine_running.id
    assert spine_running_rows["items"][0]["target"] == "catalog-feed"

    spine_stuck_rows = jobs_list(status="stuck", dataset="catalog-feed",
                                 source="spine", user=admin.username, db=db)
    assert spine_stuck_rows["total"] == 1
    assert spine_stuck_rows["items"][0]["id"] == spine_stuck.id
    assert spine_stuck_rows["items"][0]["stuck_reason"] == "heartbeat_missing_or_expired"

    spine_failed_rows = jobs_list(status="failed", dataset="catalog-feed",
                                  source="spine", user=admin.username, db=db)
    assert spine_failed_rows["total"] == 1
    assert spine_failed_rows["items"][0]["id"] == spine_failed.id
    assert spine_failed_rows["items"][0]["error"] == "fetch timeout"
    spine_detail = job_detail(spine_failed.id, source="spine",
                              user=admin.username, db=db)
    assert spine_detail["retryable"] is True

    ondemand_running_rows = jobs_list(status="running", dataset="lazada",
                                      source="ondemand", user=admin.username,
                                      db=db)
    assert ondemand_running_rows["total"] == 1
    assert ondemand_running_rows["items"][0]["id"] == ondemand_running.id

    ondemand_stuck_rows = jobs_list(status="stuck", dataset="lazada",
                                    source="ondemand", user=admin.username,
                                    db=db)
    assert ondemand_stuck_rows["total"] == 1
    assert ondemand_stuck_rows["items"][0]["id"] == ondemand_stuck.id
    assert ondemand_stuck_rows["items"][0]["stuck_reason"] == "running_timeout"

    ondemand_failed_rows = jobs_list(status="failed", dataset="shopee",
                                     source="ondemand", user=admin.username,
                                     db=db)
    assert ondemand_failed_rows["total"] == 1
    assert ondemand_failed_rows["items"][0]["id"] == ondemand_failed.id
    assert ondemand_failed_rows["items"][0]["error"] == "captcha"
    ondemand_detail = job_detail(ondemand_failed.id, source="ondemand",
                                 user=admin.username, db=db)
    assert ondemand_detail["batch_id"] == "batch-failed"
    assert ondemand_detail["retryable"] is True

    with pytest.raises(HTTPException) as exc:
        jobs_list(status="failed", source="crawl", user=alice.username, db=db)
    assert exc.value.status_code == 403


def test_workspace_jobs_list_marks_retryable_and_retry_enqueues_visible_site():
    from app.api.routes import list_jobs, retry_crawl_job

    db = _session()
    ws_a, ws_b, alice, bob, _admin = _seed_two_workspaces(db)
    failed = CrawlJob(site="site_a", status="failed",
                      requested_by_workspace_id=ws_a.id,
                      failure_code="zero_products",
                      created_at=datetime.utcnow() - timedelta(minutes=5),
                      finished_at=datetime.utcnow() - timedelta(minutes=4))
    active = CrawlJob(site="site_a", status="running",
                      requested_by_workspace_id=ws_a.id,
                      created_at=datetime.utcnow(),
                      started_at=datetime.utcnow(),
                      heartbeat_at=datetime.utcnow())
    other = CrawlJob(site="site_b", status="failed",
                     requested_by_workspace_id=ws_b.id,
                     failure_code="http_403",
                     created_at=datetime.utcnow() - timedelta(minutes=5),
                     finished_at=datetime.utcnow() - timedelta(minutes=4))
    db.add_all([failed, active, other])
    db.commit()

    out = list_jobs(limit=20, ids=f"{failed.id},{active.id},{other.id}",
                    user=alice.username,
                    x_workspace_id=str(ws_a.id), db=db)
    by_id = {row["id"]: row for row in out["items"]}
    assert by_id[failed.id]["retryable"] is True
    assert by_id[active.id]["retryable"] is False
    assert other.id not in by_id

    with patch("app.api.routes.enqueue", return_value=999) as mocked_enqueue:
        res = retry_crawl_job(failed.id, user=alice.username,
                              x_workspace_id=str(ws_a.id), db=db)

    assert res["job_id"] == 999
    assert res["retried_from"] == failed.id
    mocked_enqueue.assert_called_once_with(
        "site_a",
        trigger="admin_retry",
        requested_by_workspace_id=ws_a.id,
        requested_by_user_id=alice.id,
    )

    with pytest.raises(HTTPException) as exc:
        retry_crawl_job(other.id, user=alice.username,
                        x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 404

    with pytest.raises(HTTPException) as exc:
        retry_crawl_job(active.id, user=alice.username,
                        x_workspace_id=str(ws_a.id), db=db)
    assert exc.value.status_code == 409

    bob_out = list_jobs(limit=20, ids=f"{failed.id},{active.id},{other.id}",
                        user=bob.username,
                        x_workspace_id=str(ws_b.id), db=db)
    assert [row["id"] for row in bob_out["items"]] == [other.id]


def test_jobs_list_merges_failed_product_retry_progress_with_parent_job():
    from app.api.routes import list_jobs

    db = _session()
    ws_a, _ws_b, alice, _bob, _admin = _seed_two_workspaces(db)
    created_at = datetime(2026, 6, 28, 2, 0, 0)
    parent = CrawlJob(
        site="site_a",
        status="partial",
        trigger="scheduled",
        requested_by_workspace_id=ws_a.id,
        products_count=740,
        total_product_count=741,
        failure_code="superseded",
        created_at=created_at,
        finished_at=created_at + timedelta(hours=1),
    )
    retry = CrawlJob(
        site="site_a",
        status="success",
        trigger="failed_product_retry",
        requested_by_workspace_id=ws_a.id,
        products_count=1,
        total_product_count=1,
        created_at=created_at + timedelta(hours=6),
        finished_at=created_at + timedelta(hours=6, minutes=1),
    )
    db.add_all([parent, retry])
    db.commit()

    payload = list_jobs(
        limit=20,
        created_from="2026-06-28T00:00:00+00:00",
        created_to="2026-06-28T23:59:59+00:00",
        user=alice.username,
        x_workspace_id=str(ws_a.id),
        db=db,
    )

    assert payload["total"] == 1
    row = payload["items"][0]
    assert row["id"] == retry.id
    assert row["status"] == "success"
    assert row["products_count"] == 741
    assert row["total_product_count"] == 741
    assert row["total_product_count_source"] == "crawl_retry_merged"


def test_admin_data_quality_product_samples_are_filterable():
    from app.api.admin_spine import admin_data_quality_products

    db = _session()
    _ws_a, _ws_b, alice, _bob, admin = _seed_two_workspaces(db)
    p = db.query(Product).filter(Product.site == "site_a", Product.sku == "A-1").first()
    p.title = p.sku
    p.sale_price = None
    p.original_price = None
    p.thirty_day_sales = 0
    p.thirty_day_revenue = 0
    p.category_path = "Outdoor / Storage"
    p.status = "on_sale"
    p.created_time = datetime(2026, 6, 1, 8, 30)
    p.published_at = datetime(2026, 5, 20)
    db.add(CrawlJob(site="site_a", status="failed",
                    failure_code="parse_none",
                    suggested_action="检查解析规则",
                    created_at=datetime(2026, 6, 2)))
    db.commit()

    out = admin_data_quality_products("site_a", issue="title_weak",
                                      user=admin.username, db=db)

    assert out["site"] == "site_a"
    assert out["issue"] == "title_weak"
    assert out["total"] == 1
    assert out["issue_counts"]["all"] == 1
    assert out["issue_counts"]["title_weak"] == 1
    assert out["issue_counts"]["price_missing"] == 1
    assert [row["sku"] for row in out["items"]] == ["A-1"]
    first = out["items"][0]
    assert first["site"] == "site_a"
    assert first["category_path"] == "Outdoor / Storage"
    assert first["status"] == "on_sale"
    assert first["created_time"] == "2026-06-01T08:30:00"
    assert first["published_at"] == "2026-05-20T00:00:00"
    assert first["latest_job"]["failure_code"] == "parse_none"
    assert first["latest_job"]["suggested_action"] == "检查解析规则"
    assert "解析" in first["suggested_action"]
    assert out["items"][0]["issues"] == [
        "title_weak", "price_missing", "sales_missing", "revenue_missing",
    ]

    p2 = _product(db, "site_a", "A-WEAK-2")
    p2.title = p2.sku
    p3 = _product(db, "site_a", "A-WEAK-3")
    p3.title = p3.sku
    db.commit()
    second_page = admin_data_quality_products(
        "site_a", issue="title_weak", limit=2, page=2,
        user=admin.username, db=db)
    assert second_page["page"] == 2
    assert second_page["page_size"] == 2
    assert second_page["total"] == 3
    assert [row["sku"] for row in second_page["items"]] == ["A-1"]

    with pytest.raises(HTTPException) as exc:
        admin_data_quality_products("site_a", user=alice.username, db=db)
    assert exc.value.status_code == 403


def test_admin_data_quality_products_exposes_job_issue_details():
    from app.api.admin_spine import admin_data_quality, admin_data_quality_products

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    failed = CrawlJob(site="site_a", status="failed",
                      requested_by_workspace_id=ws_a.id,
                      failure_code="proxy_auth_failed",
                      failure_stage="fetch",
                      failure_detail="proxy auth rejected",
                      retryable=False,
                      suggested_action="修复住宅代理账号",
                      created_at=datetime.utcnow() - timedelta(minutes=10),
                      finished_at=datetime.utcnow() - timedelta(minutes=9))
    db.add(failed)
    db.commit()

    quality = admin_data_quality(user=admin.username, db=db)
    site_a = next(row for row in quality["items"] if row["site"] == "site_a")
    assert "latest_job_failed" in site_a["issues"]
    assert "proxy_auth_failed" in site_a["issues"]
    assert "proxy_unavailable" not in site_a["issues"]

    detail = admin_data_quality_products(
        "site_a", issue="proxy_auth_failed", user=admin.username, db=db)
    assert detail["kind"] == "job"
    assert detail["total"] == 1
    assert detail["issue_counts"]["proxy_auth_failed"] == 1
    assert detail["items"][0]["id"] == failed.id
    assert detail["items"][0]["failure_code"] == "proxy_auth_failed"
    assert detail["items"][0]["suggested_action"] == "修复住宅代理账号"

    stale = CrawlJob(site="site_a", status="pending",
                     requested_by_workspace_id=ws_a.id,
                     created_at=datetime.utcnow() - timedelta(hours=3))
    db.add(stale)
    db.commit()
    stale_detail = admin_data_quality_products(
        "site_a", issue="job_pending_stale", user=admin.username, db=db)
    assert stale_detail["kind"] == "job"
    assert stale_detail["total"] == 1
    assert stale_detail["items"][0]["id"] == stale.id

    partial = CrawlJob(site="site_a", status="success",
                       requested_by_workspace_id=ws_a.id,
                       failure_code="http_429",
                       failure_stage="fetch",
                       failure_detail="rate limited after partial crawl",
                       products_count=3,
                       created_at=datetime.utcnow() - timedelta(minutes=4),
                       finished_at=datetime.utcnow() - timedelta(minutes=3))
    db.add(partial)
    db.commit()
    partial_detail = admin_data_quality_products(
        "site_a", issue="partial_crawl", user=admin.username, db=db)
    assert partial_detail["kind"] == "job"
    assert partial_detail["total"] == 1
    assert partial_detail["issue_counts"]["partial_crawl"] == 1
    assert partial_detail["items"][0]["id"] == partial.id
    assert partial_detail["items"][0]["failure_code"] == "http_429"


def test_admin_data_quality_products_exposes_site_issue_details():
    from app.api.admin_spine import admin_data_quality, admin_data_quality_products

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    link = (db.query(WorkspaceSite)
            .filter(WorkspaceSite.workspace_id == ws_a.id,
                    WorkspaceSite.site == "site_a")
            .one())
    link.target_sku_count = 10
    db.commit()

    deviation = admin_data_quality_products(
        "site_a", issue="sku_deviation_high", user=admin.username, db=db)

    assert deviation["kind"] == "site"
    assert deviation["total"] == 1
    assert deviation["issue_counts"]["sku_deviation_high"] == 1
    row = deviation["items"][0]
    assert row["site"] == "site_a"
    assert row["sku_count"] == 1
    assert row["target_sku_count"] == 10
    assert row["sku_deviation_pct"] == -90.0
    assert "sku_deviation_high" in row["issues"]
    assert row["rerun_recommended"] is False
    assert row["rerun_ready"] is False
    assert row["rerun_after_setup"] is True
    assert row["rerun_blocked"] is True
    assert isinstance(row["external_data_required"], bool)

    promos = admin_data_quality_products(
        "site_a", issue="promotions_missing", user=admin.username, db=db)
    assert promos["kind"] == "site"
    assert promos["total"] == 1
    assert promos["items"][0]["promotion_count"] == 0
    assert "promotions_missing" in promos["items"][0]["issues"]

    pdp_site = _site(db, "bol_quality", "Bol")
    pdp_site.platform = "bol"
    _workspace_site(db, ws_a, "bol_quality")
    pdp_product = _product(db, "bol_quality", "BOL-NO-PRICE")
    pdp_product.sale_price = None
    pdp_product.original_price = None
    db.commit()

    quality = admin_data_quality(tenant=ws_a.id, user=admin.username, db=db)
    pdp_row = next(row for row in quality["items"] if row["site"] == "bol_quality")
    assert "price_missing" in pdp_row["issues"]
    assert "pdp_price_required" in pdp_row["issues"]
    assert "PDP" in pdp_row["suggested_action"]
    assert pdp_row["external_data_required"] is True
    assert quality["summary"]["pdp_price_required"] == 1

    pdp_detail = admin_data_quality_products(
        "bol_quality", issue="pdp_price_required", user=admin.username, db=db)
    assert pdp_detail["kind"] == "site"
    assert pdp_detail["total"] == 1
    assert pdp_detail["issue_counts"]["pdp_price_required"] == 1
    pdp_item = pdp_detail["items"][0]
    assert "pdp_price_required" in pdp_item["issues"]
    assert pdp_item["external_data_required"] is True
    assert pdp_item["rerun_recommended"] is False
    assert pdp_item["rerun_ready"] is False
    assert pdp_item["rerun_after_setup"] is True
    assert pdp_item["rerun_blocked"] is True
    assert "pdp_price_required" in pdp_item["rerun_preconditions"]
    pdp_preconditions = {
        item["issue"]: item for item in quality["summary"]["rerun_preconditions"]
    }
    assert pdp_preconditions["pdp_price_required"]["count"] == 1
    assert pdp_preconditions["pdp_price_required"]["sites"] == ["bol_quality"]

    pdp_site.crawler_config = {
        "price_source_type": "feed",
        "price_feed_url": "/tmp/bol-price-feed.csv",
        "price_feed_sku_field": "product_id",
        "price_feed_sale_price_field": "final_price",
    }
    db.add(Trend(site="bol_quality", date=date(2026, 6, 1),
                 traffic=1000, conversion_rate=2.5))
    db.commit()

    configured_quality = admin_data_quality(
        tenant=ws_a.id, user=admin.username, db=db)
    configured_row = next(
        row for row in configured_quality["items"]
        if row["site"] == "bol_quality"
    )
    assert "price_missing" in configured_row["issues"]
    assert "pdp_price_required" not in configured_row["issues"]
    assert configured_row["price_source_configured"] is True
    assert configured_row["price_source_type"] == "feed"
    assert configured_row["price_source"] == "/tmp/bol-price-feed.csv"
    assert configured_row["external_data_required"] is False
    assert configured_row["rerun_recommended"] is True
    assert configured_row["rerun_ready"] is True
    assert configured_row["rerun_blocked"] is False
    assert configured_quality["summary"]["pdp_price_required"] == 0
    assert configured_quality["summary"]["configured_price_sources"] == 1

    configured_detail = admin_data_quality_products(
        "bol_quality", issue="promotions_missing", user=admin.username, db=db)
    assert configured_detail["kind"] == "site"
    assert configured_detail["items"][0]["price_source_configured"] is True


def test_data_quality_surfaces_currency_issues_and_product_samples():
    from app.api.admin_spine import admin_data_quality_products
    from app.api.routes import data_quality

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "quality_de", "Quality")
    _workspace_site(db, ws_a, "quality_de")
    mismatch = _product(db, "quality_de", "CUR-MISMATCH")
    mismatch.title = "Currency mismatch product"
    mismatch.currency = "USD"
    mismatch.thirty_day_sales = 1
    mismatch.thirty_day_revenue = 10
    missing = _product(db, "quality_de", "CUR-MISSING")
    missing.title = "Currency missing product"
    missing.currency = None
    missing.thirty_day_sales = 1
    missing.thirty_day_revenue = 10
    db.commit()

    quality = data_quality(user=admin.username, x_workspace_id=str(ws_a.id), db=db)
    row = next(item for item in quality["items"] if item["site"] == "quality_de")

    assert row["expected_currency"] == "EUR"
    assert row["currency_missing_count"] == 1
    assert row["currency_mismatch_count"] == 1
    assert "currency_missing" in row["issues"]
    assert "currency_mismatch" in row["issues"]
    assert quality["summary"]["currency_missing"] == 1
    assert quality["summary"]["currency_mismatch"] == 1
    assert quality["summary"]["currency_issues"] == 1

    mismatch_rows = admin_data_quality_products(
        "quality_de", issue="currency_mismatch", user=admin.username, db=db)
    assert [item["sku"] for item in mismatch_rows["items"]] == ["CUR-MISMATCH"]
    assert mismatch_rows["items"][0]["currency"] == "USD"
    assert mismatch_rows["items"][0]["expected_currency"] == "EUR"
    missing_rows = admin_data_quality_products(
        "quality_de", issue="currency_missing", user=admin.username, db=db)
    assert [item["sku"] for item in missing_rows["items"]] == ["CUR-MISSING"]
    assert missing_rows["items"][0]["expected_currency"] == "EUR"


def test_data_quality_marks_non_rerunnable_market_paused():
    from app.api.admin_spine import admin_data_quality

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    _site(db, "site_paused", "P")
    _workspace_site(db, ws_a, "site_paused")
    job = CrawlJob(site="site_paused", status="failed",
                   requested_by_workspace_id=ws_a.id,
                   failure_code="market_paused",
                   failure_stage="discover",
                   failure_detail="target market is pausing orders until further notice",
                   retryable=False,
                   suggested_action="目标市场暂停运营；等待恢复",
                   created_at=datetime(2026, 6, 1),
                   finished_at=datetime(2026, 6, 1, 0, 1))
    db.add(job)
    db.flush()
    db.add(CrawlFailure(site="site_paused", job_id=job.id,
                        code="market_paused", stage="discover",
                        detail="target market is pausing orders until further notice",
                        retryable=False,
                        suggested_action="目标市场暂停运营；等待恢复",
                        occurred_at=datetime(2026, 6, 1, 0, 1)))
    db.commit()

    out = admin_data_quality(tenant=ws_a.id, user=admin.username, db=db)
    row = next(item for item in out["items"] if item["site"] == "site_paused")
    assert row["status"] == "warning"
    assert row["severity"] == "warning"
    assert row["last_error_code"] == "market_paused"
    assert row["latest_failure"]["retryable"] is False
    assert "market_paused" in row["issues"]
    assert "重跑不能解决" in row["suggested_action"]
    assert row["rerun_recommended"] is False
    assert row["rerun_blocked"] is True
    assert out["summary"]["rerun_blocked"] >= 1


def test_product_trend_is_workspace_scoped_and_returns_sales_promos():
    import asyncio
    import io
    from openpyxl import load_workbook
    from app.api.routes import export_product_trend, product_trend

    db = _session()
    ws_a, ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    site = db.query(Site).filter(Site.site == "site_a").first()
    site.review_rate = 0.05
    p = _product(db, "site_a", "TREND-SKU")
    p.sale_price = 11
    p.original_price = 15
    p.currency = "USD"
    p.title = "Trend Product Fallback"
    p.image_urls = ["https://site-a.example.com/fallback.jpg"]
    p.label = "Desk"
    p.is_new = True
    p.is_bestseller = True
    p.spu = "TREND-LISTING"
    p.review_count = 15
    p.thirty_day_sales = 100
    p.thirty_day_revenue = 1100
    db.add(Product(site="site_a", brand="A", sku="TREND-SKU-BLUE",
                   spu="TREND-LISTING", variant_id="blue",
                   title="Trend Product Blue"))
    db.add(PriceHistory(site="site_a", sku="TREND-SKU", date=date(2026, 6, 1),
                        sale_price=10, original_price=15, review_count=10))
    db.add(PriceHistory(site="site_a", sku="TREND-SKU", date=date(2026, 6, 15),
                        sale_price=11, original_price=15, review_count=15))
    db.add(Promotion(site="site_a", sku="TREND-SKU",
                     promotion_type="price_promotion",
                     promotion_name="Summer deal",
                     original_price=15, promotion_price=11,
                     discount_percent=27,
                     detected_time=datetime(2026, 6, 15)))
    db.add(Promotion(site="site_a", sku="TREND-SKU",
                     promotion_type="coupon",
                     promotion_name="Winter coupon",
                     product_title="Trend Product",
                     product_image="https://site-a.example.com/trend.jpg",
                     original_price=15, promotion_price=12,
                     discount_percent=20,
                     detected_time=datetime(2026, 6, 16, 14, 30)))
    db.commit()

    out = product_trend(pid=p.id, user=admin.username,
                        x_workspace_id=str(ws_a.id), db=db)

    assert out["product"]["sku"] == "TREND-SKU"
    assert out["summary"]["has_review_signal"] is True
    assert out["summary"]["review_rate"] == 0.05
    assert out["trend"][1]["estimated_sales"] == 100
    assert out["trend"][1]["estimated_revenue"] == 1100
    assert {p["promotion_name"] for p in out["promotions"]} == {"Summer deal", "Winter coupon"}
    summer = next(promo for promo in out["promotions"]
                  if promo["promotion_name"] == "Summer deal")
    assert summer["product_title"] == "Trend Product Fallback"
    assert summer["product_image"] == "https://site-a.example.com/fallback.jpg"
    assert summer["product_label"] == "Desk"
    assert summer["is_new"] is True
    assert summer["is_bestseller"] is True
    assert summer["listing_sku"] == "TREND-SKU"
    assert summer["variant_skus"] == ["TREND-SKU", "TREND-SKU-BLUE"]
    assert summer["variant_count"] == 2
    assert out["summary"]["current_period"]["estimated_sales"] == 100
    assert out["summary"]["promotion_total"] == 2
    assert out["summary"]["promotion_page"] == 1
    assert out["summary"]["promotion_page_size"] == 20

    page_two = product_trend(pid=p.id, user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db,
                             promo_page=2, promo_page_size=1)
    assert page_two["summary"]["promotion_total"] == 2
    assert page_two["summary"]["promotion_page"] == 2
    assert page_two["summary"]["promotion_page_size"] == 1
    assert [promo["promotion_name"] for promo in page_two["promotions"]] == ["Summer deal"]

    fallback_search = product_trend(pid=p.id, user=admin.username,
                                    x_workspace_id=str(ws_a.id), db=db,
                                    promo_search="Fallback")
    assert {promo["promotion_name"] for promo in fallback_search["promotions"]} == {"Summer deal", "Winter coupon"}

    filtered = product_trend(pid=p.id, user=admin.username,
                             x_workspace_id=str(ws_a.id), db=db,
                             granularity="month", date_from="2026-06-01",
                             date_to="2026-06-16",
                             promo_search="winter", promo_type="coupon")
    assert len(filtered["trend"]) == 1
    assert filtered["trend"][0]["date"] == "2026-06"
    assert filtered["trend"][0]["estimated_sales"] == 100
    assert [p["promotion_name"] for p in filtered["promotions"]] == ["Winter coupon"]
    assert filtered["summary"]["granularity"] == "month"
    assert filtered["summary"]["visible_points"] == 1
    response = export_product_trend(token=make_token(admin.username, ""),
                                    pid=p.id, workspace_id=ws_a.id,
                                    db=db, granularity="month",
                                    promo_search="winter")

    async def _read_body(iterator):
        chunks = []
        async for chunk in iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    content = asyncio.run(_read_body(response.body_iterator))
    wb = load_workbook(io.BytesIO(content), read_only=True)
    assert wb.sheetnames == ["Sales Trends", "Sales Promotion"]
    trend_headers = [c.value for c in next(wb["Sales Trends"].iter_rows(max_row=1))]
    promo_headers = [c.value for c in next(wb["Sales Promotion"].iter_rows(max_row=1))]
    assert trend_headers[:4] == ["Date", "Sales", "Revenue", "Ratings"]
    assert promo_headers[:4] == ["Updated Time", "SKU", "Products Details", "Product Image"]
    trend_rows = list(wb["Sales Trends"].iter_rows(min_row=2, values_only=True))
    assert trend_rows[0][5] == "USD 11"
    assert trend_rows[0][6] == "USD 15"
    promo_rows = list(wb["Sales Promotion"].iter_rows(min_row=2, values_only=True))
    assert {row[6] for row in promo_rows} == {"20%"}
    assert promo_rows[0][0] == "2026-06-16 14:30"
    assert promo_rows[0][4] == "Coupons"
    assert promo_rows[0][7] == "USD 15"
    assert promo_rows[0][8] == "USD 12"

    page_response = export_product_trend(token=make_token(admin.username, ""),
                                         pid=p.id, workspace_id=ws_a.id,
                                         db=db, export_scope="page",
                                         promo_page=2, promo_page_size=1)
    page_content = asyncio.run(_read_body(page_response.body_iterator))
    page_wb = load_workbook(io.BytesIO(page_content), read_only=True)
    page_promo_rows = list(page_wb["Sales Promotion"].iter_rows(min_row=2, values_only=True))
    assert [row[5] for row in page_promo_rows] == ["Summer deal"]

    with pytest.raises(HTTPException) as exc:
        product_trend(pid=p.id, user=admin.username,
                      x_workspace_id=str(ws_b.id), db=db)
    assert exc.value.status_code == 404


def test_product_trend_http_route_returns_payload():
    from fastapi.testclient import TestClient
    from app.db import get_db
    from app.main import app

    db = _threadsafe_session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    site = db.query(Site).filter(Site.site == "site_a").first()
    site.review_rate = 0.05
    p = _product(db, "site_a", "HTTP-TREND-SKU")
    p.sale_price = 20
    p.original_price = 30
    p.currency = "USD"
    p.review_count = 8
    p.thirty_day_sales = 7
    p.thirty_day_revenue = 140
    db.add(PriceHistory(site="site_a", sku="HTTP-TREND-SKU",
                        date=date(2026, 6, 1),
                        sale_price=19, original_price=30, review_count=5))
    db.add(PriceHistory(site="site_a", sku="HTTP-TREND-SKU",
                        date=date(2026, 6, 16),
                        sale_price=20, original_price=30, review_count=8))
    db.add(Promotion(site="site_a", sku="HTTP-TREND-SKU",
                     promotion_type="coupon",
                     promotion_name="HTTP coupon",
                     detected_time=datetime(2026, 6, 16, 10)))
    db.commit()

    def override_get_db():
        yield db

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        response = client.get(
            f"/api/products/{p.id}/trend",
            headers={
                "Authorization": f"Bearer {make_token(admin.username, '')}",
                "X-Workspace-ID": str(ws_a.id),
            },
            params={"granularity": "day", "promo_search": "HTTP"},
        )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["product"]["sku"] == "HTTP-TREND-SKU"
    assert body["summary"]["visible_points"] == 2
    assert body["summary"]["has_review_signal"] is True
    assert body["trend"][-1]["estimated_sales"] == 60
    assert body["trend"][-1]["estimated_revenue"] == 1200
    assert [row["promotion_name"] for row in body["promotions"]] == [
        "HTTP coupon",
    ]


def test_product_trend_uses_current_snapshot_without_price_history():
    from app.api.routes import product_trend

    db = _session()
    ws_a, _ws_b, _alice, _bob, admin = _seed_two_workspaces(db)
    p = db.query(Product).filter(Product.site == "site_a",
                                 Product.sku == "A-1").one()
    p.sale_price = 19.5
    p.original_price = 29.0
    p.ratings = 4.7
    p.review_count = 18
    p.thirty_day_sales = 11
    p.thirty_day_revenue = 214.5
    p.updated_time = datetime(2026, 6, 16, 9, 15)
    db.commit()

    out = product_trend(pid=p.id, user=admin.username,
                        x_workspace_id=str(ws_a.id), db=db,
                        granularity="week")

    assert len(out["trend"]) == 1
    assert out["trend"][0]["date"] == "2026-W25"
    assert out["trend"][0]["snapshot"] is True
    assert out["trend"][0]["estimated_sales"] == 11
    assert out["trend"][0]["estimated_revenue"] == 214.5
    assert out["summary"]["snapshot_fallback"] is True
    assert out["summary"]["visible_points"] == 1
    assert out["summary"]["current_period"]["review_total"] == 18
    assert "当前商品快照" in out["summary"]["data_notes"][0]


def test_workspace_admin_user_and_invite_views_are_workspace_scoped():
    from app.api.routes import (admin_create_invite, admin_list_invites,
                                admin_list_users)

    db = _session()
    ws_a, ws_b, alice, bob, _admin = _seed_two_workspaces(db)
    tenant_admin = _user(db, "tenantadmin", ws_a, role="admin")
    db.add(InviteCode(code_prefix="a", code_hash=hash_secret("invite-a"),
                      workspace_id=ws_a.id, max_uses=1, used_count=0,
                      active=True, default_role="user"))
    db.add(InviteCode(code_prefix="b", code_hash=hash_secret("invite-b"),
                      workspace_id=ws_b.id, max_uses=1, used_count=0,
                      active=True, default_role="user"))
    db.commit()

    users = admin_list_users(user=tenant_admin.username,
                             x_workspace_id=str(ws_a.id), db=db)
    assert {u["username"] for u in users} == {"alice", "admin", "tenantadmin"}
    assert "bob" not in {u["username"] for u in users}

    invites = admin_list_invites(user=tenant_admin.username,
                                 x_workspace_id=str(ws_a.id), db=db)
    assert {i["workspace_id"] for i in invites} == {ws_a.id}

    created = admin_create_invite({"workspace_id": ws_b.id},
                                  user=tenant_admin.username,
                                  x_workspace_id=str(ws_a.id),
                                  db=db)
    assert created["workspace_id"] == ws_a.id
