"""标杆追踪面板 API + 迁移测试。"""
from unittest.mock import patch

from sqlalchemy import inspect

from app.db import engine, init_db


def test_site_has_tracking_columns():
    init_db()
    cols = {c["name"] for c in inspect(engine).get_columns("sites")}
    for col in ("track_status", "source", "creator", "review_rate",
                "created_at", "updated_at"):
        assert col in cols, f"sites 缺列 {col}"


from fastapi.testclient import TestClient

from app.main import app
from app.auth import make_token


def _admin_headers():
    return {"Authorization": f"Bearer {make_token('admin', '')}"}


def test_tracking_list_requires_auth():
    init_db()
    client = TestClient(app)
    assert client.get("/api/tracking").status_code == 401


def test_tracking_list_returns_items_shape():
    init_db()
    client = TestClient(app)
    r = client.get("/api/tracking", headers=_admin_headers())
    assert r.status_code == 200
    body = r.json()
    assert "items" in body and "total" in body


def test_tracking_list_excludes_configured_hidden_sites():
    init_db()
    client = TestClient(app)
    body = client.get("/api/tracking?page_size=200", headers=_admin_headers()).json()
    listed = {item["site"] for item in body["items"]}
    assert "walmart_us" not in listed
    assert "songmics_us" in listed


def test_seeded_sites_are_marked_as_yaml_source():
    init_db()
    from app.db import SessionLocal
    from app.models import Site

    s = SessionLocal()
    try:
        songmics = s.query(Site).filter_by(site="songmics_us").one()
        assert songmics.source == "yaml"
    finally:
        s.close()


def test_add_tracking_creates_site_and_enqueues():
    init_db()
    client = TestClient(app)
    with patch("app.api.tracking.detect_platform",
               return_value=("shopify", "https://newbrand.example.com")), \
         patch("app.api.tracking.enqueue", return_value=999) as enq:
        r = client.post("/api/tracking",
                        headers=_admin_headers(),
                        json={"url": "https://newbrand.example.com/x", "brand": "NewBrand", "country": "US"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["platform"] == "shopify"
    assert body["source"] == "user"
    assert body["track_status"] == "tracking"
    enq.assert_called_once()
    assert client.get("/api/tracking", headers=_admin_headers()).json()["total"] >= 1


def test_add_tracking_falls_back_to_generic_when_undetectable():
    init_db()
    client = TestClient(app)
    with patch("app.api.tracking.detect_platform",
               return_value=(None, "https://static.example.com")), \
         patch("app.api.tracking.enqueue", return_value=1001) as enq:
        r = client.post("/api/tracking", headers=_admin_headers(),
                        json={"url": "https://static.example.com"})
    assert r.status_code == 200
    assert r.json()["platform"] == "generic"
    enq.assert_called_once()


def test_add_tracking_normalizes_long_url_and_inferrs_market_brand():
    init_db()
    client = TestClient(app)
    long_path = "/collections/desks/" + "x" * 180
    with patch("app.api.tracking.detect_platform",
               return_value=("generic", "https://flexispot.co.uk")), \
         patch("app.api.tracking.enqueue", return_value=1002):
        r = client.post("/api/tracking", headers=_admin_headers(),
                        json={"url": f"https://flexispot.co.uk{long_path}"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["url"] == "https://flexispot.co.uk"
    assert body["site"].startswith("flexispot_uk")
    assert body["country"] == "UK"
    assert body["brand"] == "flexispot"
    assert body["currency"] == "GBP"


def test_add_tracking_forbidden_for_non_admin():
    init_db()
    client = TestClient(app)
    r = client.post("/api/tracking",
                    headers={"Authorization": f"Bearer {make_token('viewer_user', '')}"},
                    json={"url": "https://x.example.com"})
    assert r.status_code in (401, 403)


def test_trigger_reuses_active_job_for_same_site():
    init_db()
    client = TestClient(app)
    code = _make_user_site(client)

    first = client.post(f"/api/jobs/trigger?site={code}", headers=_admin_headers())
    assert first.status_code == 200, first.text
    first_job = first.json()["jobs"][0]

    second = client.post(f"/api/jobs/trigger?site={code}", headers=_admin_headers())
    assert second.status_code == 200, second.text
    body = second.json()
    assert body["status"] == "already_running"
    assert body["jobs"] == [first_job]
    assert body["existing_jobs"] == [first_job]


def _make_user_site(client, *, host: str = "edit.example.com",
                    brand: str = "B", country: str = "US"):
    """借 POST 建一个 source=user 的站,返回其 site code。"""
    with patch("app.api.tracking.detect_platform",
               return_value=("shopify", f"https://{host}")), \
         patch("app.api.tracking.enqueue", return_value=1):
        r = client.post("/api/tracking", headers=_admin_headers(),
                        json={"url": f"https://{host}", "brand": brand, "country": country})
    return r.json()["site"]


def test_patch_edits_brand_and_review_rate():
    init_db(); client = TestClient(app)
    code = _make_user_site(client)
    r = client.patch(f"/api/tracking/{code}", headers=_admin_headers(),
                     json={"brand": "Edited", "review_rate": 0.03})
    assert r.status_code == 200
    assert r.json()["brand"] == "Edited"
    assert r.json()["review_rate"] == 0.03


def test_pause_and_resume():
    init_db(); client = TestClient(app)
    code = _make_user_site(client)
    assert client.post(f"/api/tracking/{code}/pause", headers=_admin_headers()).json()["track_status"] == "paused"
    assert client.post(f"/api/tracking/{code}/resume", headers=_admin_headers()).json()["track_status"] == "tracking"


def test_delete_user_site_and_remove_seed_site_from_workspace():
    init_db(); client = TestClient(app)
    code = _make_user_site(client)
    deleted = client.delete(f"/api/tracking/{code}", headers=_admin_headers())
    assert deleted.status_code == 200
    assert deleted.json()["deleted_site"] is True
    from app.db import SessionLocal
    from app.models import Site
    s = SessionLocal()
    yaml_site = s.query(Site).filter_by(source="yaml").first()
    yaml_code = yaml_site.site if yaml_site else None
    s.close()
    if yaml_code:
        removed = client.delete(f"/api/tracking/{yaml_code}", headers=_admin_headers())
        assert removed.status_code == 200
        assert removed.json()["deleted_site"] is False
        s = SessionLocal()
        try:
            assert s.query(Site).filter_by(site=yaml_code).first() is not None
        finally:
            s.close()
        listed = client.get("/api/tracking", headers=_admin_headers()).json()["items"]
        assert yaml_code not in {item["site"] for item in listed}


def test_patch_invalid_review_rate_400():
    init_db(); client = TestClient(app)
    code = _make_user_site(client)
    r = client.patch(f"/api/tracking/{code}", headers=_admin_headers(), json={"review_rate": "abc"})
    assert r.status_code == 400


def test_paused_site_skips_enqueue():
    init_db(); client = TestClient(app)
    code = _make_user_site(client)
    client.post(f"/api/tracking/{code}/pause", headers=_admin_headers())
    from app.scheduler import _product_job
    with patch("app.runner.enqueue") as enq:
        _product_job(code)
        enq.assert_not_called()


def test_export_returns_xlsx():
    init_db(); client = TestClient(app)
    _make_user_site(client)
    from app.auth import make_token
    from openpyxl import load_workbook
    import io
    r = client.get(f"/api/tracking/export?token={make_token('admin', '')}")
    assert r.status_code == 200
    assert "spreadsheet" in r.headers.get("content-type", "")
    assert r.content[:2] == b"PK"  # xlsx = zip
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    headers = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
    assert headers == ["Market", "Brand", "URL", "Status", "Products",
                       "30-Day Sales", "30-Day Revenue", "Updated Time",
                       "Created Time", "Creator", "Last Error"]
    assert "Created Time" in headers
    assert "Creator" in headers
    assert "30-Day Revenue" in headers
    assert client.get("/api/tracking/export?token=bogus").status_code == 401


def test_tracking_filters_metrics_latest_job_and_export_scope():
    init_db(); client = TestClient(app)
    from datetime import datetime
    from openpyxl import load_workbook
    import io
    import uuid
    from app.db import SessionLocal
    from app.models import CrawlJob, Product

    suffix = uuid.uuid4().hex[:8]
    host = f"filter-metric-{suffix}.example.com"
    brand = f"MetricBrand{suffix}"
    code = _make_user_site(client, host=host, brand=brand, country="CA")
    s = SessionLocal()
    try:
        s.add(Product(site=code, sku=f"METRIC-{suffix}", spu=f"SPU-{suffix}",
                      title="Metric Product", sale_price=20,
                      thirty_day_sales=3, thirty_day_revenue=60,
                      updated_time=datetime(2026, 6, 16)))
        s.add(CrawlJob(site=code, status="failed", trigger="manual",
                       failure_code="http_403", failure_stage="fetch",
                       failure_detail="HTTP 403 forbidden",
                       suggested_action="配置可用住宅代理",
                       created_at=datetime(2026, 6, 16, 1),
                       finished_at=datetime(2026, 6, 16, 1, 1)))
        s.commit()
    finally:
        s.close()

    listed = client.get(
        f"/api/tracking?search={suffix}&market=CA&brand={brand}&status=tracking",
        headers=_admin_headers(),
    )
    assert listed.status_code == 200, listed.text
    body = listed.json()
    assert body["total"] == 1
    row = body["items"][0]
    assert row["site"] == code
    assert row["products"] == 1
    assert row["sku_count"] == 1
    assert row["thirty_day_sales"] == 3
    assert row["thirty_day_revenue"] == 60
    assert row["sales_available"] is True
    assert row["revenue_available"] is True
    assert row["track_status"] == "tracking"
    assert row["display_status"] == "error"
    assert row["last_error_code"] == "http_403"
    assert row["latest_job"]["failure_stage"] == "fetch"

    error_listed = client.get(
        f"/api/tracking?search={suffix}&market=CA&brand={brand}&status=error",
        headers=_admin_headers(),
    )
    assert error_listed.status_code == 200, error_listed.text
    error_body = error_listed.json()
    assert error_body["total"] == 1
    assert error_body["items"][0]["site"] == code
    assert "error" in error_body["facets"]["statuses"]

    token = make_token("admin", "")
    exported = client.get(
        f"/api/tracking/export?token={token}&search={suffix}&market=CA&brand={brand}&status=tracking"
    )
    assert exported.status_code == 200, exported.text
    wb = load_workbook(io.BytesIO(exported.content), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert len(rows) == 2
    assert rows[1][0:5] == ("CA", brand, f"https://{host}", "error", 1)
    assert rows[1][5] == 3
    assert rows[1][7] == "2026-06-16 00:00"
    assert "T" not in rows[1][8]
    assert rows[1][10] == "http_403"


def test_tracking_revenue_does_not_depend_on_sales_signal():
    init_db(); client = TestClient(app)
    from datetime import datetime
    from openpyxl import load_workbook
    import io
    import uuid
    from app.db import SessionLocal
    from app.models import Product

    suffix = uuid.uuid4().hex[:8]
    host = f"revenue-only-{suffix}.example.com"
    brand = f"RevenueOnly{suffix}"
    code = _make_user_site(client, host=host, brand=brand, country="US")
    s = SessionLocal()
    try:
        s.add(Product(site=code, sku=f"REV-{suffix}", spu=f"REVSPU-{suffix}",
                      title="Revenue Only Product", sale_price=25,
                      thirty_day_sales=0, thirty_day_revenue=125,
                      updated_time=datetime(2026, 6, 15)))
        s.commit()
    finally:
        s.close()

    listed = client.get(f"/api/tracking?search={suffix}",
                        headers=_admin_headers())
    assert listed.status_code == 200, listed.text
    row = listed.json()["items"][0]
    assert row["thirty_day_sales"] == 0
    assert row["thirty_day_revenue"] == 125
    assert row["sales_available"] is False
    assert row["revenue_available"] is True

    token = make_token("admin", "")
    exported = client.get(f"/api/tracking/export?token={token}&search={suffix}")
    assert exported.status_code == 200, exported.text
    wb = load_workbook(io.BytesIO(exported.content), read_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[1][5] in ("", None)
    assert rows[1][6] == "USD 125.0"
