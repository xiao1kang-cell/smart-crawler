#!/usr/bin/env python3
"""Build a status audit from the benchmark acceptance workbook.

The workbook still contains the original acceptance marks, so this script does
not rewrite it.  It creates a separate Markdown audit that maps each open row to
the current implementation evidence and the remaining closure step.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/Users/wangxiaokang/Desktop/标杆平台验收报告.xlsx")
DEFAULT_OUTPUT = Path("docs/acceptance-status-audit.md")


@dataclass(frozen=True)
class IssueRow:
    sheet: str
    row: int
    module: str
    area: str
    description: str
    workbook_status: str
    note: str


@dataclass(frozen=True)
class Rule:
    status: str
    owner: str
    evidence: str
    patterns: tuple[str, ...]

    def matches(self, text: str) -> bool:
        return any(re.search(pattern, text, re.I) for pattern in self.patterns)


RULES: tuple[Rule, ...] = (
    Rule(
        "已代码覆盖",
        "权限",
        "前台 /report 和 /app 路由要求登录，报表编辑/导出/自定义动作由 canEdit 按角色门控。",
        (
            r"权限",
            r"仅有权限",
        ),
    ),
    Rule(
        "已代码覆盖",
        "前端/接口",
        "SiteReportPage + report/export API 已实现产品分析筛选、列表、导出、趋势和促销明细。",
        (
            r"Store Analysis|产品分析|Product Analysis|趋势图|BestSelling|Newest",
            r"Category|销售促销|Sales Promotion|刷新|导出|操作选型|日期筛选|类型筛选",
            r"Product Trend|By Month|By Week|By Days|时间维度",
            r"销售趋势板块里面的数据不可修改，不可查看",
        ),
    ),
    Rule(
        "已代码覆盖",
        "前端/接口",
        "TrackingPage + tracking API 已实现新增、搜索筛选、状态、销售收入、时间、创建人、操作和分页。",
        (
            r"Add Tracking|新增标杆|搜索框|Market|Brand|Status|国旗|URL",
            r"Updated Time|Created Time|Creator",
            r"Stop Tracking|Edit|Delete|分页栏",
        ),
    ),
    Rule(
        "待外部数据导入",
        "数据",
        "流量/转化率不是页面抓取稳定产物，后台已提供第三方指标导入/校验入口。",
        (
            r"流量",
            r"转化率|conversion",
        ),
    ),
    Rule(
        "待真实数据重跑验证",
        "数据",
        "后台数据质量页已暴露缺价格、缺促销、SKU 偏差、任务失败和重跑前置条件；关闭需要生产抓取结果。",
        (
            r"偏差|商品数量|34个网站|16个网站",
            r"SKU数量|SKU数据|SKU 数据|SKU 数据不一致",
            r"商品名称没有抓取完整|货币符号|促销数据",
            r"产品列表.*(Sales Price|Price|Sales|Revenues)",
            r"汇总数据.*(30-Day Sales|30-Day Revenues|Price|SKU)",
            r"趋势图.*(Revenues|Price)",
            r"30-Day Sales|30-Day Revenue|30天销量|30天收入",
            r"爬取过来的竞品数据没有价格",
            r"报表数据不全|绝大部分网站|少数网站",
            r"数据好像也不对",
        ),
    ),
    Rule(
        "待浏览器验收",
        "前端",
        "相关样式已收敛到页面组件，但还需要用真实页面截图做视觉验收。",
        (
            r"深色模式|文字对比",
            r"样式|不和谐",
        ),
    ),
    Rule(
        "已代码覆盖",
        "后台",
        "QueuePage + admin jobs API 已按真实队列表聚合，并暴露运行中、失败、完成时间和详情。",
        (
            r"采集任务|任务列表|完成日期|运行中|失败|52成功|只显示40|队列",
        ),
    ),
    Rule(
        "已代码覆盖",
        "后台/代理",
        "ProxiesPage + proxy endpoints/pools/rules API 已区分普通 IP、住宅 IP、代理池和站点规则。",
        (
            r"代理池|住宅代理|普通ip|住宅ip|SOCKS5|HTTP",
        ),
    ),
    Rule(
        "已代码覆盖",
        "后台",
        "DataQualityPage + data-quality API 已提供站点级和明细级问题清单。",
        (
            r"能看到是哪些数据|明细|数据质量|人肉",
        ),
    ),
    Rule(
        "待浏览器验收",
        "验收",
        "这是规格覆盖率的总述，需要用本对账表和真实页面逐项复核后关闭。",
        (
            r"120条功能.*27条",
        ),
    ),
)


def _cell(value: object) -> str:
    return str(value).strip() if value is not None else ""


def load_issues(workbook_path: Path) -> list[IssueRow]:
    wb = load_workbook(workbook_path, data_only=True)
    rows: list[IssueRow] = []

    ws = wb["功能点检查"]
    current_module = ""
    current_area = ""
    for idx in range(2, ws.max_row + 1):
        module = _cell(ws.cell(idx, 1).value) or current_module
        area = _cell(ws.cell(idx, 2).value) or current_area
        desc = _cell(ws.cell(idx, 3).value)
        status = _cell(ws.cell(idx, 4).value)
        note = _cell(ws.cell(idx, 6).value)
        if module:
            current_module = module
        if area:
            current_area = area
        if not desc:
            continue
        if status == "×" or note:
            rows.append(IssueRow(ws.title, idx, module, area, desc, status, note))

    ws = wb["问题"]
    for idx in range(2, ws.max_row + 1):
        desc = _cell(ws.cell(idx, 1).value)
        if desc:
            rows.append(IssueRow(ws.title, idx, "验收问题", "", desc, "问题", ""))

    ws = wb["爬取数据"]
    header = [_cell(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    for idx in range(2, ws.max_row + 1):
        values = [_cell(ws.cell(idx, col).value) for col in range(1, ws.max_column + 1)]
        if not any(values):
            continue
        site = values[0]
        deviation = values[3] if len(values) > 3 else ""
        if site and deviation:
            desc = " / ".join(
                f"{h}:{v}" for h, v in zip(header[:4], values[:4]) if h and v
            )
            rows.append(IssueRow(ws.title, idx, "爬取数据", site, desc, "数据", ""))

    return rows


def classify(row: IssueRow) -> Rule:
    text = " ".join(
        part for part in (row.module, row.area, row.description, row.note) if part
    )
    for rule in RULES:
        if rule.matches(text):
            return rule
    return Rule(
        "需继续确认",
        "产品/研发",
        "脚本没有找到明确实现证据，需要人工确认功能边界或补充规则。",
        tuple(),
    )


def _short(text: str, limit: int = 92) -> str:
    text = re.sub(r"\s+", " ", text).strip()
    return text if len(text) <= limit else text[: limit - 1] + "..."


def _summary(rows: Iterable[tuple[IssueRow, Rule]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for _, rule in rows:
        counts[rule.status] = counts.get(rule.status, 0) + 1
    return counts


def render_markdown(workbook_path: Path, classified: list[tuple[IssueRow, Rule]]) -> str:
    counts = _summary(classified)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    unresolved = [
        item for item in classified
        if item[1].status != "已代码覆盖"
    ]
    lines = [
        "# 标杆平台验收状态对账",
        "",
        f"- 生成时间: {now}",
        f"- 验收表: `{workbook_path}`",
        f"- 问题行总数: {len(classified)}",
        "",
        "## 状态汇总",
        "",
        "| 状态 | 数量 | 含义 |",
        "| --- | ---: | --- |",
    ]
    meanings = {
        "已代码覆盖": "本地代码已有对应页面/API/测试入口，仍建议浏览器或接口抽验。",
        "待真实数据重跑验证": "能力已具备，但必须跑生产抓取任务并看结果才能关闭。",
        "待外部数据导入": "需要 SimilarWeb/GA/BI/人工文件等第三方指标，重跑抓取本身无法生成。",
        "待浏览器验收": "需要用真实页面做视觉/交互验收。",
        "需继续确认": "脚本未找到明确证据，需补需求或补实现。",
    }
    for status, count in sorted(counts.items(), key=lambda item: item[0]):
        lines.append(f"| {status} | {count} | {meanings.get(status, '')} |")

    lines.extend([
        "",
        "## 仍未关闭",
        "",
        "| Sheet:行 | 模块 | 问题 | 当前状态 | 下一步 |",
        "| --- | --- | --- | --- | --- |",
    ])
    for row, rule in unresolved:
        lines.append(
            f"| {row.sheet}:{row.row} | {_short(row.area or row.module, 30)} | "
            f"{_short(row.description)} | {rule.status} | {rule.evidence} |"
        )

    lines.extend([
        "",
        "## 已找到代码覆盖的原验收行",
        "",
        "| Sheet:行 | 模块 | 问题 | 证据 |",
        "| --- | --- | --- | --- |",
    ])
    for row, rule in classified:
        if rule.status == "已代码覆盖":
            lines.append(
                f"| {row.sheet}:{row.row} | {_short(row.area or row.module, 30)} | "
                f"{_short(row.description)} | {rule.evidence} |"
            )

    lines.extend([
        "",
        "## 关闭口径",
        "",
        "- 功能类: 代码覆盖后，还需要至少一次前端构建和真实页面点验。",
        "- 数据类: 必须在生产/准生产环境重跑对应站点后，以后台数据质量页和站点报表为准。",
        "- 外部指标类: 流量、转化率必须先导入第三方指标，再刷新报表。",
        "- 代理类: 代理池不是只看配置数量，必须看健康检查、站点规则命中和失败分类。",
    ])
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    issues = load_issues(args.workbook)
    classified = [(row, classify(row)) for row in issues]
    markdown = render_markdown(args.workbook, classified)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(markdown, encoding="utf-8")

    counts = _summary(classified)
    print(f"wrote {args.output}")
    print(f"issues={len(classified)}")
    for status, count in sorted(counts.items(), key=lambda item: item[0]):
        print(f"{status}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
